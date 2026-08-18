#!/usr/bin/env python3
"""UNIFIED internal-control figure: UPR-sensor transcripts by BRANCH, bulk + snRNA together, one
colour per branch (matching the manuscript panels i/r): EIF2AK3/PERK = cyan, ERN1/IRE1 = maroon,
ATF6 = blue.

3 rows = branches. Columns = contexts: bulk [Thapsigargin, Mizuno, Nativio] | snRNA [Ex, In, Ast,
Mic, Oli, OPC]. Y = log2 fold-change of each replicate vs the control/low reference group mean
(bulk: disease sample / control mean ; snRNA: late-Braak donor / Braak I-II mean). Each point = one
sample (bulk) or one donor (snRNA); bar = mean; dashed line = no change; two-sided Mann-Whitney U
(disease/late vs control/low). Shows sensor transcripts do NOT change -> they are internal reference
genes, not activation readouts (Reviewer 2, comment 1)."""
import os, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams.update({"font.family": "Arial", "font.sans-serif": ["Arial"], "axes.unicode_minus": False})
import openpyxl
from scipy.stats import mannwhitneyu

ROOT = "/data/adbrainseq"
BASE = os.path.join(ROOT, "Publication/2024_scRNA seq/V1 Science manuscript/ADBrainSeq_V6.0/"
                    "Acta Neuropathologica Communication")
OUT = os.path.join(BASE, "Major Revision", "processed raw data")
BRANCHES = [("EIF2AK3", "PERK", "#5FC7D8", "#1F8EA0"),
            ("ERN1", "IRE1", "#9C2A4E", "#611026"),
            ("ATF6", "ATF6", "#3E52A0", "#232E63")]
GSET = [b[0] for b in BRANCHES]

# ================= BULK per-sample log2FC vs control mean =================
def thap(fname):
    p = os.path.join(ROOT, "Stanford U/PERK BioSensor Bulk RNA Seq Analysis/Data analysis/"
                     "1st DATA from BGI/GEO deposit/geo_submission_Nov/processed data", fname)
    wb = openpyxl.load_workbook(p, read_only=True); ws = wb.active
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    fc = [i for i, c in enumerate(hdr) if c and "fpkm" in str(c).lower()]
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        s = str(r[2]).strip() if r[2] else ""
        if s in GSET: out[s] = [float(r[i]) for i in fc if isinstance(r[i], (int, float))]
    wb.close(); return out
thap_c, thap_t = thap("RNA seq_Control FPKM.xlsx"), thap("RNA seq_Thapsigargin FPKM.xlsx")

mf = os.path.join(ROOT, "Stanford U/PERK Human AD brain analysis/Human AD brain SEQ analysis/"
                  "Alzheimer's brain disease Bulk RNA seq/2021 Mizuno_Human AD brain RNA seq_decreased PERK/"
                  "2021 Mizuno_GSE173955_Table_1_gene_level_expression.xlsx")
wb = openpyxl.load_workbook(mf, read_only=True); ws = wb["2021 Mizuno_DESeq"]
miz_ad, miz_c = {}, {}
for r in ws.iter_rows(min_row=4, values_only=True):
    s = str(r[0]).strip() if r[0] else ""
    if s in GSET:
        miz_ad[s] = [float(x) for x in r[1:9] if isinstance(x, (int, float))]
        miz_c[s] = [float(x) for x in r[12:22] if isinstance(x, (int, float))]
wb.close()

wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD3_NativioSeq.xlsx"), read_only=True); ws = wb.active
tot_ad = np.zeros(12); tot_old = np.zeros(10); nat_raw = {}
for r in ws.iter_rows(min_row=3, values_only=True):
    s = str(r[1]).strip() if r[1] else ""
    ad = [x if isinstance(x, (int, float)) else 0.0 for x in r[6:18]]
    old = [x if isinstance(x, (int, float)) else 0.0 for x in r[18:28]]
    if len(ad) == 12 and len(old) == 10:
        tot_ad += np.array(ad, float); tot_old += np.array(old, float)
        if s in GSET: nat_raw[s] = (np.array(ad, float), np.array(old, float))
wb.close()
nat_ad, nat_c = {}, {}
for s, (ad, old) in nat_raw.items():
    nat_ad[s] = list(ad / tot_ad * 1e6); nat_c[s] = list(old / tot_old * 1e6)

def log2fc_samples(dis, ctrl):
    """per-sample log2FC vs the control GEOMETRIC mean (log-space ref; robust to outliers)."""
    cl = [math.log2(v) for v in ctrl if v > 0]
    if not cl: return np.array([]), np.array([])
    ref = float(np.mean(cl))                                  # log2(geomean_control)
    d = [math.log2(v) - ref for v in dis if v > 0]
    c = [math.log2(v) - ref for v in ctrl if v > 0]
    return np.array(d), np.array(c)

# ================= snRNA donor-level log2FC (late vs low) from npy =================
rows = np.load(os.path.join(OUT, "R2Q1_snRNA_sensor_donormeans.npy"), allow_pickle=True)
sn = {g: {} for g in GSET}                    # sn[gene][ct] = {'low':{'fc':[],'lin':[],'und':0}, 'late':{...}}
for r in rows:
    g, ct, grp, v, fc = r[0], r[1], r[4], float(r[5]), r[6]
    if g not in sn or grp not in ("low", "late"): continue
    d = sn[g].setdefault(ct, {}).setdefault(grp, {"fc": [], "lin": [], "und": 0})
    d["lin"].append(v)
    if fc is not None and np.isfinite(float(fc)): d["fc"].append(float(fc))
    else: d["und"] += 1
SN_CELLS = [("Ex", "Ex"), ("In", "In"), ("Ast", "Ast"), ("Mic", "Mic"), ("Oli", "OligD"), ("Opc", "OPC")]

# ================= assemble columns per gene =================
def contexts(gene):
    """returns list of dicts: {label, fc(detected disease log2FC), und(undetected count),
    td/tc (values for the disease-vs-control Mann-Whitney test)}"""
    out = []
    for lab, dis, ctrl in [("Thapsigargin", thap_t.get(gene, []), thap_c.get(gene, [])),
                           ("Mizuno", miz_ad.get(gene, []), miz_c.get(gene, [])),
                           ("Nativio", nat_ad.get(gene, []), nat_c.get(gene, []))]:
        fc, _ = log2fc_samples(dis, ctrl)
        out.append({"label": lab, "fc": fc, "und": 0,
                    "td": np.array([v for v in dis if v > 0]), "tc": np.array([v for v in ctrl if v > 0])})
    for ct, lab in SN_CELLS:
        late = sn[gene].get(ct, {}).get("late", {"fc": [], "lin": [], "und": 0})
        low = sn[gene].get(ct, {}).get("low", {"fc": [], "lin": [], "und": 0})
        out.append({"label": lab, "fc": np.array(late["fc"]), "und": late["und"],
                    "td": np.array(late["lin"]), "tc": np.array(low["lin"])})
    return out

def pstar(p):
    return "****" if p < 1e-4 else "***" if p < 1e-3 else "**" if p < 1e-2 else "*" if p < 0.05 else "ns"

NCOL = 9
fig, axes = plt.subplots(3, 1, figsize=(13.5, 9.4))
YLIM = (-2.9, 2.6)
FLOOR = YLIM[0] + 0.20
for gi, (gene, alias, fill, edge) in enumerate(BRANCHES):
    ax = axes[gi]
    ctx = contexts(gene)
    ax.axhspan(-0.585, 0.585, color="#f2f2f2", zorder=0)
    ax.axhline(0, color="#888", lw=0.9, ls="--", zorder=1)
    ax.axhline(FLOOR + 0.13, color="#e2e2e2", lw=0.6, zorder=0)
    ax.axvline(2.5, color="#bbb", lw=1.2, ls=":", zorder=1)          # bulk | snRNA divider
    for xi, cx in enumerate(ctx):
        d = cx["fc"]
        if len(d) >= 2:
            parts = ax.violinplot([d], positions=[xi], widths=0.72, showmeans=False, showextrema=False)
            for pc in parts["bodies"]:
                pc.set_facecolor(fill); pc.set_alpha(0.45); pc.set_edgecolor(edge); pc.set_linewidth(1.1)
        if len(d):
            jit = (np.random.RandomState(gi*30+xi).rand(len(d)) - 0.5) * 0.22
            ax.scatter(xi + jit, d, s=14, color=fill, edgecolor="white", linewidth=0.4, zorder=3)
            ax.hlines(d.mean(), xi-0.26, xi+0.26, color=edge, lw=1.8, zorder=4)
        if cx["und"]:                                                # undetected -> hollow floor markers
            jit = (np.random.RandomState(77+gi*30+xi).rand(cx["und"]) - 0.5) * 0.22
            ax.scatter(xi + jit, np.full(cx["und"], FLOOR), s=13, facecolors="none",
                       edgecolors=edge, linewidth=0.8, zorder=3)
        if len(cx["td"]) >= 2 and len(cx["tc"]) >= 2:
            p = mannwhitneyu(cx["td"], cx["tc"], alternative="two-sided").pvalue
            ax.text(xi, YLIM[1]-0.28, pstar(p), ha="center", va="top", fontsize=7.5, color="#333")
    ax.set_ylim(*YLIM); ax.set_xlim(-.6, NCOL-0.4)
    ax.set_xticks(range(NCOL))
    ax.set_xticklabels([f"{cx['label']}\n(n={len(cx['fc'])+cx['und']})" for cx in ctx], fontsize=7.6)
    ax.tick_params(labelsize=8, length=0)
    for sp in ("top", "right"): ax.spines[sp].set_visible(False)
    ax.set_ylabel(f"{gene} / {alias}\nlog$_2$FC vs control", fontsize=9.5, fontweight="bold", color=edge)
    if gi == 0:
        ax.text(1.0, YLIM[1]+0.15, "BULK  (disease vs control)", ha="center", va="bottom",
                fontsize=9, fontweight="bold", color="#555")
        ax.text(6.0, YLIM[1]+0.15, "snRNA  (Braak V–VI vs I–II, donor-level)", ha="center", va="bottom",
                fontsize=9, fontweight="bold", color="#555")

fig.suptitle("UPR-sensor transcripts do not change with disease — bulk and snRNA, by branch",
             fontsize=13, fontweight="bold", y=0.997)
fig.text(0.5, 0.958, "Each point = one sample (bulk: Thapsigargin 5v5, Mizuno 8 AD vs 10 non-AD, Nativio 12 AD vs 10 old) "
         "or one donor (snRNA late-Braak, n=17) · log$_2$FC vs the control/low-Braak group mean · bar = mean · "
         "grey band = ±1.5-fold · two-sided Mann–Whitney U.",
         ha="center", va="top", fontsize=8, color="#555")
fig.subplots_adjust(top=0.92, bottom=0.055, left=0.085, right=0.99, hspace=0.42)
for ext in ("png", "svg"):
    fig.savefig(os.path.join(OUT, f"R2Q1_sensor_unified_branch.{ext}"),
                dpi=300 if ext == "png" else None, facecolor="white", bbox_inches="tight")
print("saved R2Q1_sensor_unified_branch.{png,svg}")
for gene, alias, *_ in BRANCHES:
    for cx in contexts(gene):
        d = cx["fc"]
        p = (mannwhitneyu(cx["td"], cx["tc"], alternative="two-sided").pvalue
             if len(cx["td"]) >= 2 and len(cx["tc"]) >= 2 else float("nan"))
        m = f"{d.mean():+.2f}" if len(d) else "  NA"
        print(f"  {gene:8s} {cx['label']:12s} n={len(d)+cx['und']:2d} (und={cx['und']}) meanLog2FC={m}  MWU p={p:.2g}")
