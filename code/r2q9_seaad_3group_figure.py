#!/usr/bin/env python3
"""SEA-AD (independent) — Braak III,IV / I,II  and  Braak V,VI / I,II  (Braak 0,I,II = control),
manuscript Figure-4 style (two-tone green violins, red mean, blue quartiles). Per cell type two violins:
per-gene log2FC of the Response-to-ER-stress set for III/IV-vs-control and V/VI-vs-control. Significance
is DONOR-LEVEL (Mann-Whitney on the per-donor UPR-set score vs the 6 control donors)."""
import os, csv, statistics as st
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams.update({"font.family": "Arial", "font.sans-serif": ["Arial"], "axes.unicode_minus": False, "pdf.fonttype": 42, "ps.fonttype": 42})
from matplotlib.patches import Patch
from scipy.stats import mannwhitneyu, wilcoxon

OUT = os.path.dirname(os.path.abspath(__file__))
PB = os.path.join(OUT, "SEAAD_3group_donor_pseudobulk.csv")
SENS = {"EIF2AK3", "ERN1", "ATF6"}
CELLS = [("Ex", "Excitatory\nneurons"), ("In", "Inhibitory\nneurons"), ("Ast", "Astrocytes"),
         ("Mic", "Microglia"), ("Oli", "Oligodendrocytes"), ("OPC", "Oligodendrocyte\nprogenitor cells")]
C_MID, C_HIGH = "#3c8f34", "#c8d94a"        # (legacy) Fig-4 two-tone; superseded by per-cell-type CTCOL below
RED, BLUE = "#e2231a", "#2b3a8f"
# Figure-5 style: colour each cell type by its own hue (both Braak groups share the hue); regular labels, no stats.
CTCOL = {"Ex": "#746fbd", "In": "#c2b7de", "Ast": "#f2a9a0",
         "Mic": "#cdea9a", "Oli": "#9ad4e7", "OPC": "#cbb0e0"}

rows = list(csv.DictReader(open(PB)))
genecols = [g for g in rows[0].keys() if g not in ("cell_type", "donor", "braak_group")]
tgt = [g for g in genecols if g not in SENS]
# per cell type: group -> list of per-donor gene vectors
byct = {c: {"low": [], "mid": [], "high": []} for c, _ in CELLS}
for r in rows:
    c = r["cell_type"]
    if c not in byct: continue
    byct[c][r["braak_group"]].append(np.array([float(r[g]) for g in tgt]))
# per-donor MODULE SCORE = mean of z-scored (across all donors within the cell type) UPR-gene expression
# (robust to baseline; same style as the discovery-cohort R2.2 donor module score)
score = {c: {"low": [], "mid": [], "high": []} for c, _ in CELLS}
for c, _ in CELLS:
    order = [(g, "low") for g in byct[c]["low"]] + [(g, "mid") for g in byct[c]["mid"]] + [(g, "high") for g in byct[c]["high"]]
    M = np.vstack([v for v, _ in order]); mu = M.mean(0); sd = M.std(0); sd[sd == 0] = 1
    Z = (M - mu) / sd
    for k, (_, grp) in enumerate(order):
        score[c][grp].append(float(Z[k].mean()))

# ---- ordinal-Braak trend (donor-level, the reviewer-preferred test): Spearman(module score, Braak 0-6) over ALL donors ----
import openpyxl
from scipy.stats import spearmanr
wb = openpyxl.load_workbook(os.path.join(OUT, "SEAAD_donor_metadata_SuppTable1.xlsx"), read_only=True); ws = wb.active
h = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
di = next(i for i, c in enumerate(h) if "donor id" in c.lower()); bi = next(i for i, c in enumerate(h) if c.lower().startswith("braak"))
bmap = {"Braak 0": 0, "Braak I": 1, "Braak II": 2, "Braak III": 3, "Braak IV": 4, "Braak V": 5, "Braak VI": 6}
d2b = {r[di]: bmap.get(str(r[bi])) for r in ws.iter_rows(min_row=2, values_only=True) if r[di]}; wb.close()
trend = {}
for c, _ in CELLS:
    dd = [r["donor"] for r in rows if r["cell_type"] == c]
    M = np.vstack([np.array([float(r[g]) for g in tgt]) for r in rows if r["cell_type"] == c])
    mu = M.mean(0); sd = M.std(0); sd[sd == 0] = 1; sc = ((M - mu) / sd).mean(1)
    bk = np.array([d2b.get(d, np.nan) for d in dd]); ok = ~np.isnan(bk)
    trend[c] = spearmanr(sc[ok], bk[ok])

def pstar(p): return "****" if p < 1e-4 else "***" if p < 1e-3 else "**" if p < 1e-2 else "*" if p < 0.05 else "ns"

# per-gene log2FC vs low mean
def gene_fc(c, grp):
    lo = np.vstack(byct[c]["low"]).mean(0); g = np.vstack(byct[c][grp]).mean(0)
    return g - lo

lim = 0
for c, _ in CELLS:
    for grp in ("mid", "high"):
        lim = max(lim, np.abs(gene_fc(c, grp)).max())
lim *= 1.12
fig, ax = plt.subplots(figsize=(12.5, 5.8))
ax.axhline(0, color="#000", lw=0.8, ls=":", zorder=1)
YL = (-lim, lim); bw = 0.34
xt, xtl = [], []
for i, (c, lab) in enumerate(CELLS):
    for k, (grp, glab) in enumerate([("mid", "III,IV"), ("high", "V,VI")]):
        xpos = i + (k - 0.5) * bw
        v = np.clip(gene_fc(c, grp), *YL)
        parts = ax.violinplot([v], positions=[xpos], widths=bw * 0.92, showmeans=False, showextrema=False)
        for pc in parts["bodies"]:
            pc.set_facecolor(CTCOL[c]); pc.set_alpha(0.62); pc.set_edgecolor("#1f1f1f"); pc.set_linewidth(0.7)
        jit = (np.random.RandomState(i * 2 + k).rand(len(v)) - 0.5) * bw * 0.7
        ax.scatter(xpos + jit, v, s=3, color="#111", alpha=0.45, edgecolor="none", zorder=3)
        q1, q3 = np.percentile(v, [25, 75]); ax.hlines([q1, q3], xpos - bw * .45, xpos + bw * .45, color=BLUE, lw=1.1, zorder=5)
        ax.hlines(v.mean(), xpos - bw * .48, xpos + bw * .48, color=RED, lw=2.4, zorder=6)
        xt.append(xpos); xtl.append(glab)
ax.set_xlim(-0.6, len(CELLS) - 0.4); ax.set_ylim(*YL)
ax.set_xticks(xt); ax.set_xticklabels(xtl, fontsize=8)
# cell-type headers below the Braak-group ticks (regular weight, Fig-5 style)
from matplotlib.transforms import blended_transform_factory
tr = blended_transform_factory(ax.transData, ax.transAxes)
for i, (c, lab) in enumerate(CELLS):
    ax.text(i, -0.105, lab, ha="center", va="top", fontsize=9.5, transform=tr, linespacing=0.95)
ax.set_ylabel("log$_2$FC  early-AD or late-AD / low-Braak\n(Response to ER stress, genes=254)", fontsize=10)
for sp in ("top", "right"): ax.spines[sp].set_visible(False)
ax.legend(handles=[plt.Line2D([0], [0], color=RED, lw=2.4, label="set mean"),
                   plt.Line2D([0], [0], color=BLUE, lw=1.1, label="quartiles")],
          frameon=False, fontsize=7.5, loc="lower right", ncol=2)
# titles/subtitles removed -> moved to the Figure 5 legend (publication style)
fig.subplots_adjust(top=0.95, bottom=0.17, left=0.09, right=0.985)
for ext in ("png", "svg", "pdf"):
    fig.savefig(os.path.join(OUT, f"R2Q9_SEAAD_3group_figure.{ext}"), dpi=300 if ext == "png" else None,
                facecolor="white", bbox_inches="tight")
print("saved R2Q9_SEAAD_3group_figure.{png,svg}")
for c, _ in CELLS:
    for grp in ("mid", "high"):
        fc = gene_fc(c, grp); p = mannwhitneyu(score[c][grp], score[c]["low"], alternative="two-sided").pvalue
        print(f"  {c:4s} {grp:4s} vs low: gene-log2FC mean={fc.mean():+.4f} {100*(fc<0).sum()//len(fc)}% down | "
              f"donor {np.mean(score[c][grp]):+.3f}(n={len(score[c][grp])}) vs {np.mean(score[c]['low']):+.3f}(n={len(score[c]['low'])}) p={p:.2g}")
