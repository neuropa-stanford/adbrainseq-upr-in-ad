#!/usr/bin/env python3
"""Excel: mRNA transport (internal control) gene list with log2FC in Mizuno, Nativio (CPM),
and snRNA-seq per cell type."""
import sys, os, math, statistics
sys.path.insert(0, '.')
from r1q1_gomatrix import load_mizuno
from nativio_cpm import load_nativio_cpm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = ("/data/adbrainseq/Publication/2024_scRNA seq/"
        "V1 Science manuscript/ADBrainSeq_V6.0/Acta Neuropathologica Communication")
OUT = os.path.join(BASE, "Major Revision", "processed raw data")

genes = [l.strip() for l in open(os.path.join(OUT, "controlset_MRNA_TRANSPORT.txt")) if l.strip()]
miz = {g: v[0] for g, v in load_mizuno().items()}
nat = load_nativio_cpm()

# snRNA per-cell-type log2FC from SuppD6
wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD6_snRNSeqDB.xlsx"), read_only=True); ws = wb.active
CT_E = {"Ex": 1, "In": 3, "Ast": 5, "Mic": 7, "OligD": 9, "OPC": 11}  # early Braak
CT_L = {"Ex": 2, "In": 4, "Ast": 6, "Mic": 8, "OligD": 10, "OPC": 12}  # late Braak
ALLCT = [("E", CT_E), ("L", CT_L)]
sn = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    d = {}
    for stg, ctm in ALLCT:
        for ct, ci in ctm.items():
            try:
                fc = float(r[ci])
                if fc > 0 and math.isfinite(fc): d[f"{ct}_{stg}"] = math.log2(fc)
            except (TypeError, ValueError): pass
    sn[str(r[0]).strip()] = d
wb.close()

wb = openpyxl.Workbook(); wsx = wb.active; wsx.title = "mRNA transport control"
thin = Side(style="thin", color="C8C8C8"); border = Border(*[thin] * 4)
center = Alignment(horizontal="center", vertical="center")
BULK = PatternFill("solid", fgColor="EAF1F7"); SNR = PatternFill("solid", fgColor="EAF3EA")
BULKH = PatternFill("solid", fgColor="C9DDEC"); SNRH = PatternFill("solid", fgColor="CDE5CD")

# title
wsx.merge_cells("A1:O1")
wsx["A1"] = "Internal-control gene set: mRNA transport (GOBP_MRNA_TRANSPORT) — log2 fold-change (AD vs control)"
wsx["A1"].font = Font(bold=True, size=12); wsx["A1"].alignment = center
# group headers
wsx.merge_cells("B2:C2"); wsx["B2"] = "bulk RNA-seq"; wsx["B2"].font = Font(bold=True, color="12507E")
wsx["B2"].alignment = center; wsx["B2"].fill = BULKH
wsx.merge_cells("D2:I2"); wsx["D2"] = "snRNA-seq — EARLY Braak (III/IV vs I/II)"; wsx["D2"].font = Font(bold=True, color="1d5a38")
wsx["D2"].alignment = center; wsx["D2"].fill = SNRH
wsx.merge_cells("J2:O2"); wsx["J2"] = "snRNA-seq — LATE Braak (V/VI vs I/II)"; wsx["J2"].font = Font(bold=True, color="8c5a1d")
wsx["J2"].alignment = center; wsx["J2"].fill = PatternFill("solid", fgColor="F0E4CE")
# column headers
cts = ["Ex", "In", "Ast", "Mic", "OligD", "OPC"]
hdr = ["Gene", "Mizuno", "Nativio\n(CPM)"] + cts + cts
LATEH = PatternFill("solid", fgColor="F0E4CE")
for j, h in enumerate(hdr):
    c = wsx.cell(row=3, column=j + 1, value=h); c.font = Font(bold=True); c.alignment = center; c.border = border
    c.fill = BULKH if 1 <= j <= 2 else (SNRH if 3 <= j <= 8 else (LATEH if j >= 9 else PatternFill("solid", fgColor="ECECEC")))

def bval(d, g):
    return round(d[g], 3) if g in d else None
def sval(g, key):
    return round(sn[g][key], 3) if g in sn and key in sn[g] else None
LATE = PatternFill("solid", fgColor="FBF3E4")
r = 4
for g in sorted(genes):
    row = [g, bval(miz, g), bval(nat, g)] + [sval(g, f"{ct}_E") for ct in cts] + [sval(g, f"{ct}_L") for ct in cts]
    for j, v in enumerate(row):
        c = wsx.cell(row=r, column=j + 1, value=v); c.border = border; c.alignment = center
        if j == 0: c.font = Font(bold=True)
        elif 1 <= j <= 2: c.fill = BULK
        elif 3 <= j <= 8: c.fill = SNR
        else: c.fill = LATE
    r += 1
# median row
def med(vs):
    vs = [v for v in vs if v is not None]; return round(statistics.median(vs), 3) if vs else None
cols = list(zip(*[[bval(miz, g), bval(nat, g)] + [sval(g, f"{ct}_E") for ct in cts] + [sval(g, f"{ct}_L") for ct in cts] for g in genes]))
wsx.cell(row=r, column=1, value="median").font = Font(bold=True, italic=True)
for j, cv in enumerate(cols):
    c = wsx.cell(row=r, column=j + 2, value=med(cv)); c.font = Font(bold=True, italic=True); c.alignment = center
    c.fill = BULKH if j <= 1 else (SNRH if j <= 7 else LATEH); c.border = border

wsx.column_dimensions["A"].width = 12
for col in "BCDEFGHIJKLMNO": wsx.column_dimensions[col].width = 8.5
wsx.row_dimensions[3].height = 30; wsx.freeze_panes = "B4"
wsx.auto_filter.ref = f"A3:O{r-1}"
wb.save(os.path.join(OUT, "mRNA_transport_control_expression.xlsx"))
print(f"saved mRNA_transport_control_expression.xlsx ({len(genes)} genes)")
print("EARLY medians: " + str([med(c) for c in cols[2:8]]) + "  LATE: " + str([med(c) for c in cols[8:]]))
print("Miz %s Nat %s | earlyEx %s In %s Ast %s Mic %s Oli %s OPC %s"
      % tuple(med(c) for c in cols[:8]))
