#!/usr/bin/env python3
"""Internal-control sensor-gene expression: EIF2AK3(PERK), ERN1(IRE1), ATF6.
Per-sample violins (n = each sample) in the three bulk datasets:
  Thapsigargin (in-vitro positive control): Control(fpkm_B,n=5) vs Thapsigargin(fpkm_F,n=5) [FPKM]
  Mizuno GSE173955: non-AD(n=10) vs AD(n=8) [expression level, as provided]
  Nativio:          Old(n=10)   vs AD(n=12) [raw counts -> CPM, log2]
Group identity verified: Mizuno via GSE173955 clinical metadata (RNA sample sheet);
Nativio via column sample labels (-AD / -Old); Thap via GEO processed file names.
Two-sided Mann-Whitney U p per gene/dataset; individual points overlaid."""
import os, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams.update({"font.family": "Arial", "font.sans-serif": ["Arial"], "axes.unicode_minus": False})
import openpyxl
from scipy.stats import mannwhitneyu

ROOT = "/data/adbrainseq"
BASE = os.path.join(ROOT, "Publication/2024_scRNA seq/V1 Science manuscript/ADBrainSeq_V6.0/"
                    "Acta Neuropathologica Communication")
OUT = os.path.join(BASE, "Major Revision", "processed raw data")
GENES = [("EIF2AK3", "PERK"), ("ERN1", "IRE1"), ("ATF6", "ATF6")]
GSET = [g for g, _ in GENES]

# ---------- Thapsigargin (FPKM) ----------
def load_thap_fpkm(fname):
    p = os.path.join(ROOT, "Stanford U/PERK BioSensor Bulk RNA Seq Analysis/Data analysis/"
                     "1st DATA from BGI/GEO deposit/geo_submission_Nov/processed data", fname)
    wb = openpyxl.load_workbook(p, read_only=True); ws = wb.active
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    fcols = [i for i, c in enumerate(hdr) if c and "fpkm" in str(c).lower()]
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        s = str(r[2]).strip() if r[2] else ""
        if s in GSET:
            out[s] = [float(r[i]) for i in fcols if isinstance(r[i], (int, float))]
    wb.close(); return out
thap_ctrl = load_thap_fpkm("RNA seq_Control FPKM.xlsx")
thap_treat = load_thap_fpkm("RNA seq_Thapsigargin FPKM.xlsx")

# ---------- Mizuno (expression level, DESeq sheet) ----------
mf = os.path.join(ROOT, "Stanford U/PERK Human AD brain analysis/Human AD brain SEQ analysis/"
                  "Alzheimer's brain disease Bulk RNA seq/2021 Mizuno_Human AD brain RNA seq_decreased PERK/"
                  "2021 Mizuno_GSE173955_Table_1_gene_level_expression.xlsx")
wb = openpyxl.load_workbook(mf, read_only=True); ws = wb["2021 Mizuno_DESeq"]
miz_ad, miz_ctrl = {}, {}
for r in ws.iter_rows(min_row=4, values_only=True):
    s = str(r[0]).strip() if r[0] else ""
    if s in GSET:
        miz_ad[s] = [float(x) for x in r[1:9] if isinstance(x, (int, float))]      # 8 AD
        miz_ctrl[s] = [float(x) for x in r[12:22] if isinstance(x, (int, float))]  # 10 non-AD
wb.close()

# ---------- Nativio (counts -> CPM, log2) ----------
wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD3_NativioSeq.xlsx"), read_only=True); ws = wb.active
nat_rows = []
tot_ad = np.zeros(12); tot_old = np.zeros(10)
for r in ws.iter_rows(min_row=3, values_only=True):
    s = str(r[1]).strip() if r[1] else ""
    ad = [x if isinstance(x, (int, float)) else 0.0 for x in r[6:18]]
    old = [x if isinstance(x, (int, float)) else 0.0 for x in r[18:28]]
    if len(ad) == 12 and len(old) == 10:
        tot_ad += np.array(ad, float); tot_old += np.array(old, float)
        if s in GSET: nat_rows.append((s, np.array(ad, float), np.array(old, float)))
wb.close()
nat_ad, nat_old = {}, {}
for s, ad, old in nat_rows:
    nat_ad[s] = list(np.log2(ad / tot_ad * 1e6 + 1))
    nat_old[s] = list(np.log2(old / tot_old * 1e6 + 1))

# ---------- assemble ----------
# per dataset: (name, unit, ctrl_dict, disease_dict, ctrl_label, dis_label, ctrl_color, dis_color)
DATASETS = [
    ("Thapsigargin", "FPKM", thap_ctrl, thap_treat, "Control", "Thap", "#b8b8b8", "#8e44ad"),
    ("Mizuno (GSE173955)", "expression level", miz_ctrl, miz_ad, "non-AD", "AD", "#b8b8b8", "#c0392b"),
    ("Nativio", "log$_2$ CPM", nat_old, nat_ad, "Old", "AD", "#b8b8b8", "#2c7fb8"),
]

def pstar(p):
    return "****" if p < 1e-4 else "***" if p < 1e-3 else "**" if p < 1e-2 else "*" if p < 0.05 else "ns"

fig, axes = plt.subplots(len(GENES), len(DATASETS), figsize=(10.2, 8.4))
for gi, (gene, alias) in enumerate(GENES):
    for di, (dname, unit, cd, dd, clab, dlab, cc, dcol) in enumerate(DATASETS):
        ax = axes[gi, di]
        c = np.array(cd.get(gene, []), float); d = np.array(dd.get(gene, []), float)
        parts = ax.violinplot([c, d], positions=[0, 1], widths=0.8, showmeans=False,
                              showextrema=False)
        for pc, col in zip(parts["bodies"], [cc, dcol]):
            pc.set_facecolor(col); pc.set_alpha(0.35); pc.set_edgecolor(col); pc.set_linewidth(1.2)
        for pos, arr, col in [(0, c, cc), (1, d, dcol)]:
            jit = (np.random.RandomState(gi * 10 + di).rand(len(arr)) - 0.5) * 0.22
            ax.scatter(pos + jit, arr, s=16, color=col, edgecolor="white", linewidth=0.5, zorder=3)
            ax.hlines(np.mean(arr), pos - 0.24, pos + 0.24, color="#222", lw=1.6, zorder=4)
        try:
            p = mannwhitneyu(c, d, alternative="two-sided").pvalue
        except ValueError:
            p = float("nan")
        ymax = max(c.max() if len(c) else 0, d.max() if len(d) else 0)
        ymin = min(c.min() if len(c) else 0, d.min() if len(d) else 0)
        rng = (ymax - ymin) or 1.0
        ax.plot([0, 0, 1, 1], [ymax + .05 * rng, ymax + .10 * rng, ymax + .10 * rng, ymax + .05 * rng],
                lw=1.0, color="#444")
        ax.text(0.5, ymax + .12 * rng, f"{pstar(p)}\np={p:.2g}", ha="center", va="bottom", fontsize=7.2)
        ax.set_ylim(ymin - .12 * rng, ymax + .34 * rng)
        ax.set_xlim(-.6, 1.6)
        ax.set_xticks([0, 1]); ax.set_xticklabels([f"{clab}\n(n={len(c)})", f"{dlab}\n(n={len(d)})"], fontsize=8)
        ax.tick_params(labelsize=7.5, length=0)
        for sp in ("top", "right"): ax.spines[sp].set_visible(False)
        if di == 0:
            ax.set_ylabel(f"{gene} / {alias}\n\nexpression", fontsize=9, fontweight="bold")
        else:
            ax.set_ylabel("")
        ax.annotate(unit, xy=(0.02, 0.98), xycoords="axes fraction", fontsize=6.6, color="#888",
                    va="top", ha="left")
        if gi == 0:
            ax.set_title(dname, fontsize=10, fontweight="bold", pad=8)

fig.suptitle("UPR sensor transcript levels across datasets (per-sample; internal reference genes)",
             fontsize=12, fontweight="bold", y=0.995)
fig.text(0.5, 0.965, "EIF2AK3/PERK · ERN1/IRE1 · ATF6 — each point = one sample; bar = mean; "
         "two-sided Mann–Whitney U. Note: sensor mRNA level is not a measure of sensor activation (Reviewer 2, comment 1).",
         ha="center", fontsize=8, color="#555")
fig.subplots_adjust(top=0.93, bottom=0.06, left=0.10, right=0.985, hspace=0.42, wspace=0.30)
for ext in ("png", "svg"):
    fig.savefig(os.path.join(OUT, f"R2Q1_sensor_expression_bulk_violins.{ext}"),
                dpi=300 if ext == "png" else None, facecolor="white", bbox_inches="tight")
print("saved R2Q1_sensor_expression_bulk_violins.{png,svg}")
# print numbers
for gene, alias in GENES:
    for dname, unit, cd, dd, clab, dlab, *_ in DATASETS:
        c = np.array(cd.get(gene, []), float); d = np.array(dd.get(gene, []), float)
        p = mannwhitneyu(c, d, alternative="two-sided").pvalue if len(c) and len(d) else float("nan")
        print(f"  {gene:8s} {dname:18s} {clab} mean={c.mean():.2f} (n={len(c)}) | "
              f"{dlab} mean={d.mean():.2f} (n={len(d)})  MWU p={p:.3g}")
