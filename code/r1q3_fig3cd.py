#!/usr/bin/env python3
"""Reproduce manuscript Figure 3c (tSNE by broad.cell.type, split by Braak group) and 3d (cell-type
proportion % per Braak group, mean +/- SEM across donors), from the local Mathys 2019 metadata.
Same colour scheme as the manuscript. No Sherlock; all local. No pandas (numpy2/pandas ABI clash)."""
import os, csv, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams.update({"font.family": "Arial", "font.sans-serif": ["Arial"], "axes.unicode_minus": False, "pdf.fonttype": 42, "ps.fonttype": 42})
from matplotlib.gridspec import GridSpec
from matplotlib.lines import Line2D
import openpyxl

MATH = ("/data/adbrainseq/Stanford U/Collaboration support/"
        "Prof. Eun-hye Joe Ajou/2019 Mathys/Gene Expression (RNA seq)/filtered_column_metadata.txt")
CLIN = ("/data/adbrainseq/Stanford U/PERK Seth Genetics Project/"
        "ROSMAP_clinical_PMI included for Seth.xlsx")
OUT = ("/data/adbrainseq/Publication/2024_scRNA seq/V1 Science manuscript/"
       "ADBrainSeq_V6.0/Acta Neuropathologica Communication/Major Revision/processed raw data")

# cell-type display order, labels, colours (manuscript palette + Per added)
CT = ["Ex", "In", "Ast", "Mic", "Oli", "Opc", "End", "Per"]
LAB = {"Ex": "Excitatory neurons (Ex)", "In": "Inhibitory neurons (In)", "Ast": "Astrocytes (Ast)",
       "Mic": "Microglia (Mic)", "Oli": "Oligodendrocytes (OligD)", "Opc": "Oligodendrocyte progenitor cells (OPC)",
       "End": "Endothelial cells (End)", "Per": "Pericytes (Per)"}
SHORT = {"Ex": "Ex", "In": "In", "Ast": "Ast", "Mic": "Mic", "Oli": "OligD", "Opc": "OPC", "End": "End", "Per": "Per"}
COL = {"Ex": "#8CBF43", "In": "#33A45C", "Ast": "#F0806A", "Mic": "#28C2D4", "Oli": "#4F86C6",
       "Opc": "#9C77B4", "End": "#E3A43A", "Per": "#B94FA0"}
BK = [("low", "Braak I,II", "#EE3524"), ("int", "Braak III,IV", "#3B6DB3"), ("late", "Braak V,VI", "#54B948")]

# ---- projid -> braaksc ----
wb = openpyxl.load_workbook(CLIN, read_only=True); ws = wb.active
h = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
pjc, bkc = h.index("projid"), h.index("braaksc")
pb = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    try: pb[str(int(r[pjc]))] = int(r[bkc])
    except (TypeError, ValueError): pass
wb.close()
def grp(b): return "low" if b <= 2 else ("int" if b <= 4 else "late")

# ---- read metadata ----
x, y, ct, gp, proj = [], [], [], [], []
with open(MATH) as f:
    rd = csv.DictReader(f, delimiter="\t")
    for r in rd:
        p = r["projid"]
        if p not in pb: continue
        try: xx, yy = float(r["tsne1"]), float(r["tsne2"])
        except (TypeError, ValueError): continue
        c = r["broad.cell.type"]
        x.append(xx); y.append(yy); ct.append(c); gp.append(grp(pb[p])); proj.append(p)
x = np.array(x); y = np.array(y); ct = np.array(ct); gp = np.array(gp); proj = np.array(proj)
ndon = {g: len(set(proj[gp == g])) for g, _, _ in BK}
print("donors per group:", ndon, " cells:", len(x))

# ---- 3d proportions: per donor, % of each cell type; mean +/- SEM per Braak group ----
prop = {g: {c: [] for c in CT} for g, _, _ in BK}
for d in sorted(set(proj)):
    m = proj == d; tot = int(m.sum()); g = gp[m][0]
    for c in CT:
        prop[g][c].append(100.0 * int((ct[m] == c).sum()) / tot)
mean = {g: [np.mean(prop[g][c]) for c in CT] for g, _, _ in BK}
sem = {g: [np.std(prop[g][c], ddof=1) / math.sqrt(len(prop[g][c])) for c in CT] for g, _, _ in BK}

# ================= figure =================
fig = plt.figure(figsize=(13, 10))
gs = GridSpec(2, 3, height_ratios=[1.05, 0.9], hspace=0.32, wspace=0.08,
              left=0.07, right=0.72, top=0.95, bottom=0.08)
# --- 3c: tSNE, one panel per Braak group ---
PLOT_ORDER = ["Ex", "Oli", "In", "Ast", "Opc", "Mic", "End", "Per"]   # big first, small on top
for j, (g, glab, _) in enumerate(BK):
    ax = fig.add_subplot(gs[0, j])
    sel = gp == g
    for c in PLOT_ORDER:
        mm = sel & (ct == c)
        ax.scatter(x[mm], y[mm], s=2.0, c=COL[c], linewidths=0, rasterized=True)
    ax.set_title(f"{glab} (n={ndon[g]})", fontsize=14.5, bbox=dict(facecolor="#e6e6e6", edgecolor="none", pad=3))
    ax.set_xlim(-62, 56); ax.set_ylim(-60, 62); ax.set_xticks([-60, -30, 0, 30, 60])
    ax.set_xlabel("tsne1", fontsize=14.5)
    if j == 0: ax.set_ylabel("tsne2", fontsize=14.5); ax.set_yticks([-60, -30, 0, 30, 60])
    else: ax.set_yticklabels([])
    ax.tick_params(labelsize=14.5)
    ax.text(-0.16 if j == 0 else -0.05, 1.06, "C" if j == 0 else "", transform=ax.transAxes,
            fontsize=26.5, fontweight="bold")
# tSNE legend (right)
leg = [Line2D([0], [0], marker="o", ls="", mfc=COL[c], mec="none", ms=8, label=LAB[c]) for c in CT]
fig.legend(handles=leg, loc="center left", bbox_to_anchor=(0.735, 0.76), frameon=False,
           fontsize=14.5, title="broad.cell.type", title_fontsize=14.5, handletextpad=0.4, labelspacing=0.7)

# --- 3d: grouped bar of proportions ---
axd = fig.add_subplot(gs[1, :])
nX = len(CT); bw = 0.26
xpos = np.arange(nX)
for k, (g, glab, col) in enumerate(BK):
    axd.bar(xpos + (k-1)*bw, mean[g], bw, yerr=sem[g], color=col, label=glab,
            error_kw=dict(elinewidth=1, capsize=2.5, capthick=1), edgecolor="none")
axd.set_xticks(xpos); axd.set_xticklabels([SHORT[c] for c in CT], fontsize=14.5)
axd.set_ylabel("Proportion %", fontsize=14.5); axd.tick_params(labelsize=14.5)
for sp in ("top", "right"): axd.spines[sp].set_visible(False)
axd.legend(frameon=False, fontsize=14.5, loc="upper right")
axd.text(-0.06, 1.04, "D", transform=axd.transAxes, fontsize=26.5, fontweight="bold")

for ext in ("png", "svg", "pdf"):
    fig.savefig(os.path.join(OUT, f"R1Q3_Figure3cd_reproduced.{ext}"),
                dpi=300 if ext == "png" else None, facecolor="white", bbox_inches="tight")
print("saved R1Q3_Figure3cd_reproduced.{png,svg}")
print("\n3d proportions (mean%):")
for g, glab, _ in BK:
    print(f"  {glab:14s} " + "  ".join(f"{SHORT[c]}={mean[g][i]:.1f}" for i, c in enumerate(CT)))
