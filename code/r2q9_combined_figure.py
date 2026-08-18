#!/usr/bin/env python3
"""COMBINED independent-replication figure — SEA-AD (top) + Mathys 2024 (bottom), side by side,
manuscript Figure-4 style. Per cell type, two violins = per-gene log2FC of the Response-to-ER-stress
set for Braak III,IV vs 0,I,II control (dark) and Braak V,VI vs 0,I,II control (light).
Top stars per violin = gene-level Wilcoxon signed-rank (same test as manuscript Fig 4).
Bottom (orange) = DONOR-level ordinal-Braak trend (Spearman of per-donor z-scored UPR module vs Braak 0-6).

Braak source per cohort:
  SEA-AD  : per-donor Braak joined from SEAAD_donor_metadata_SuppTable1.xlsx (its CSV has no braak_num).
  Mathys24: per-donor Braak read directly from the 'braak_num' column of its pseudobulk CSV.
Draws only the cohorts whose CSV is present (so it runs with SEA-AD alone until Mathys lands)."""
import os, csv
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams.update({"font.family": "Arial", "font.sans-serif": ["Arial"], "axes.unicode_minus": False})
from matplotlib.patches import Patch
from scipy.stats import wilcoxon, spearmanr

OUT = os.path.dirname(os.path.abspath(__file__))
SENS = {"EIF2AK3", "ERN1", "ATF6"}
META_EXCL = {"cell_type", "donor", "braak_group", "braak_num"}
CELLS = [("Ex", "Excitatory\nneurons"), ("In", "Inhibitory\nneurons"), ("Ast", "Astrocytes"),
         ("Mic", "Microglia"), ("Oli", "Oligodendrocytes"), ("OPC", "Oligodendrocyte\nprogenitor cells")]
C_MID, C_HIGH = "#3c8f34", "#c8d94a"        # Fig-4 two-tone: III/IV dark green, V/VI light yellow-green
RED, BLUE = "#e2231a", "#2b3a8f"

def pstar(p):
    if p != p: return "ns"
    return "****" if p < 1e-4 else "***" if p < 1e-3 else "**" if p < 1e-2 else "*" if p < 0.05 else "ns"

def load_seaad_braak():
    """donor -> Braak 0-6 from SEA-AD Supplementary Table 1."""
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(OUT, "SEAAD_donor_metadata_SuppTable1.xlsx"), read_only=True); ws = wb.active
    h = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    di = next(i for i, c in enumerate(h) if "donor id" in c.lower())
    bi = next(i for i, c in enumerate(h) if c.lower().startswith("braak"))
    bmap = {"Braak 0": 0, "Braak I": 1, "Braak II": 2, "Braak III": 3, "Braak IV": 4, "Braak V": 5, "Braak VI": 6}
    d2b = {r[di]: bmap.get(str(r[bi])) for r in ws.iter_rows(min_row=2, values_only=True) if r[di]}
    wb.close(); return d2b

def read_cohort(pb_csv, braak_from_col):
    """Return (byct, tgt_genes, donor_braak_per_celltype). braak_from_col=True -> use braak_num column."""
    rows = list(csv.DictReader(open(pb_csv)))
    genecols = [g for g in rows[0].keys() if g not in META_EXCL]
    tgt = [g for g in genecols if g not in SENS]
    d2b = None if braak_from_col else load_seaad_braak()
    byct = {c: {"low": [], "mid": [], "high": []} for c, _ in CELLS}
    # per cell type: list of (donor, braak_num, vector) for the donor-level trend
    dl = {c: [] for c, _ in CELLS}
    for r in rows:
        c = r["cell_type"]
        if c not in byct: continue
        vec = np.array([float(r[g]) for g in tgt])
        byct[c][r["braak_group"]].append(vec)
        if braak_from_col:
            bn = int(r["braak_num"]) if r.get("braak_num", "-1") not in ("", "-1", "nan") else None
        else:
            bn = d2b.get(r["donor"])
        dl[c].append((r["donor"], bn, vec))
    return byct, tgt, dl

def gene_fc(byct, c, grp):
    lo = np.vstack(byct[c]["low"]).mean(0); g = np.vstack(byct[c][grp]).mean(0)
    return g - lo

def donor_trend(dl, c):
    """Spearman(per-donor z-scored UPR module score, Braak 0-6) over all donors with a Braak value."""
    items = [(bn, vec) for (_, bn, vec) in dl[c] if bn is not None]
    if len(items) < 4: return (np.nan, np.nan)
    M = np.vstack([v for _, v in items]); mu = M.mean(0); sd = M.std(0); sd[sd == 0] = 1
    sc = ((M - mu) / sd).mean(1); bk = np.array([bn for bn, _ in items])
    return spearmanr(sc, bk)

def draw_panel(ax, byct, dl, title, ngene, ndonor):
    lim = 0
    for c, _ in CELLS:
        for grp in ("mid", "high"):
            if byct[c][grp] and byct[c]["low"]:
                lim = max(lim, np.abs(gene_fc(byct, c, grp)).max())
    lim = (lim or 0.1) * 1.12
    YL = (-lim, lim); bw = 0.34
    ax.axhline(0, color="#000", lw=0.8, ls=":", zorder=1)
    for i, (c, lab) in enumerate(CELLS):
        for k, (grp, col) in enumerate([("mid", C_MID), ("high", C_HIGH)]):
            if not byct[c][grp] or not byct[c]["low"]: continue
            xpos = i + (k - 0.5) * bw
            fc = gene_fc(byct, c, grp); v = np.clip(fc, *YL)
            parts = ax.violinplot([v], positions=[xpos], widths=bw * 0.92, showmeans=False, showextrema=False)
            for pc in parts["bodies"]:
                pc.set_facecolor(col); pc.set_alpha(0.92); pc.set_edgecolor("#1f1f1f"); pc.set_linewidth(0.7)
            jit = (np.random.RandomState(i * 2 + k).rand(len(v)) - 0.5) * bw * 0.7
            ax.scatter(xpos + jit, v, s=2.5, color="#111", alpha=0.42, edgecolor="none", zorder=3)
            q1, q3 = np.percentile(v, [25, 75]); ax.hlines([q1, q3], xpos - bw * .45, xpos + bw * .45, color=BLUE, lw=1.0, zorder=5)
            ax.hlines(v.mean(), xpos - bw * .48, xpos + bw * .48, color=RED, lw=2.2, zorder=6)
            pw = wilcoxon(fc).pvalue
            ax.text(xpos, YL[1] * 0.86, pstar(pw), ha="center", fontsize=8.5, fontweight="bold")
        rho, pp = donor_trend(dl, c)
        ax.text(i, YL[0] * 0.82, f"r={rho:+.2f}\np={pp:.2g} {pstar(pp)}", ha="center",
                fontsize=6.0, color="#7a4a00", linespacing=1.0, fontweight="bold")
    ax.set_xlim(-0.6, len(CELLS) - 0.4); ax.set_ylim(*YL)
    ax.set_xticks(range(len(CELLS))); ax.set_xticklabels([l for _, l in CELLS], fontsize=9)
    ax.set_ylabel(f"log$_2$FC vs Braak 0,I,II\n(ER-stress, genes={ngene})", fontsize=9.5)
    for sp in ("top", "right"): ax.spines[sp].set_visible(False)
    ax.set_title(title, fontsize=11.5, fontweight="bold", loc="left", pad=6)

# ---- load whichever cohorts are present ----
panels = []
sea = os.path.join(OUT, "SEAAD_3group_donor_pseudobulk.csv")
mat = os.path.join(OUT, "MATHYS24_Braak56vs012_donor_pseudobulk.csv")
if os.path.exists(sea):
    byct, tgt, dl = read_cohort(sea, braak_from_col=False)
    nd = len({d for c in dl for (d, _, _) in dl[c]})
    panels.append(("SEA-AD  (Allen · MTG · independent · " + str(nd) + " donors)", byct, dl, len(tgt)))
if os.path.exists(mat):
    byct, tgt, dl = read_cohort(mat, braak_from_col=True)
    nd = len({d for c in dl for (d, _, _) in dl[c]})
    panels.append(("Mathys 2024  (ROSMAP · PFC · " + str(nd) + " donors)", byct, dl, len(tgt)))

n = len(panels)
fig, axes = plt.subplots(n, 1, figsize=(12.5, 4.6 * n), squeeze=False)
for ax, (title, byct, dl, ngene) in zip(axes[:, 0], panels):
    draw_panel(ax, byct, dl, title, ngene, None)
axes[0, 0].legend(handles=[Patch(facecolor=C_MID, edgecolor="#1f1f1f", label="early-AD (Braak III,IV) vs non-AD"),
                           Patch(facecolor=C_HIGH, edgecolor="#1f1f1f", label="late-AD (Braak V,VI) vs non-AD"),
                           plt.Line2D([0], [0], color=RED, lw=2.2, label="set mean"),
                           plt.Line2D([0], [0], color=BLUE, lw=1.0, label="quartiles")],
                  frameon=False, fontsize=7.5, loc="lower right", ncol=2)
fig.suptitle("UPR-associated transcription vs AD-stage (ROSMAP Braak) by cell type — two cohorts",
             fontsize=13, fontweight="bold", y=0.995)
fig.text(0.5, 0.965 if n == 2 else 0.95,
         "Groups by ROSMAP Braak: non-AD (0,I,II) = control · early-AD (III,IV, dark) · late-AD (V,VI, light).  Violins = per-gene log2FC.  "
         "Top stars = gene-level Wilcoxon (Fig-4 test).  Orange = DONOR-level ordinal-Braak trend (Spearman, pseudoreplication-free).", ha="center", fontsize=7.4, color="#555")
fig.subplots_adjust(top=0.93 if n == 2 else 0.86, bottom=0.07, left=0.09, right=0.985, hspace=0.32)
for ext in ("png", "svg"):
    fig.savefig(os.path.join(OUT, f"R2Q9_combined_figure.{ext}"), dpi=300 if ext == "png" else None,
                facecolor="white", bbox_inches="tight")
print(f"saved R2Q9_combined_figure.{{png,svg}} with {n} panel(s):", [p[0] for p in panels])
