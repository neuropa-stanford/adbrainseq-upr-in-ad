#!/usr/bin/env python3
"""Editable Excel version of the R1.5 reproducibility table."""
import csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = ("/data/adbrainseq/Publication/2024_scRNA seq/"
       "V1 Science manuscript/ADBrainSeq_V6.0/Acta Neuropathologica Communication/"
       "Major Revision/processed raw data")

rows = list(csv.reader(open(os.path.join(OUT, "R1Q5_reproducibility_table.csv"))))
header = rows[0]; data = rows[1:]

wb = openpyxl.Workbook(); ws = wb.active; ws.title = "R1.5 reproducibility"
thin = Side(style="thin", color="C8C8C8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

DOWN_C = PatternFill("solid", fgColor="EAF1F7"); DOWN_R = PatternFill("solid", fgColor="D5E6F0")
UP_C = PatternFill("solid", fgColor="FBF0EF");   UP_R = PatternFill("solid", fgColor="F6DDD9")
DOWN_H = PatternFill("solid", fgColor="DFE6EC");  UP_H = PatternFill("solid", fgColor="EFE1DF")
GS_H = PatternFill("solid", fgColor="ECECEC")

# Row 1: title
ws.merge_cells("A1:K1")
ws["A1"] = "Cross-modality directional reproducibility of UPR gene sets in AD brain transcriptomics analysis"
ws["A1"].font = Font(bold=True, size=13); ws["A1"].alignment = center

# Row 2: group headers
ws.merge_cells("B2:F2"); ws["B2"] = "DOWN-regulated"
ws["B2"].font = Font(bold=True, size=12, color="12507E"); ws["B2"].alignment = center; ws["B2"].fill = DOWN_H
ws.merge_cells("G2:K2"); ws["G2"] = "UP-regulated"
ws["G2"].font = Font(bold=True, size=12, color="9A2B1E"); ws["G2"].alignment = center; ws["G2"].fill = UP_H

# Row 3: column headers
for j, h in enumerate(header):
    c = ws.cell(row=3, column=j + 1, value=h)
    c.font = Font(bold=True, size=10, color="1B2A3A"); c.alignment = center; c.border = border
    c.fill = GS_H if j == 0 else (DOWN_H if j <= 5 else UP_H)

# Data rows
for i, row in enumerate(data):
    for j, val in enumerate(row):
        try: val = int(val)
        except (ValueError, TypeError): pass
        c = ws.cell(row=4 + i, column=j + 1, value=val)
        c.alignment = center; c.border = border
        if j == 0:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFFFFF")
        elif j in (1, 2, 3, 4): c.fill = DOWN_C
        elif j == 5: c.fill = DOWN_R; c.font = Font(bold=True, color="12507E")
        elif j in (6, 7, 8, 9): c.fill = UP_C
        elif j == 10: c.fill = UP_R; c.font = Font(bold=True, color="9A2B1E")

# footnote
fn = 4 + len(data) + 1
ws.merge_cells(start_row=fn, start_column=1, end_row=fn, end_column=11)
ws.cell(row=fn, column=1,
        value=("Bulk overlapped = genes changed in the same direction in both bulk cohorts (Mizuno and "
               "Nativio). Reproduced in snRNA = of those, the genes changed in the same direction in "
               "single-nucleus RNA-seq (neurons for down, glia for up); % is of the bulk-overlapped set."))
ws.cell(row=fn, column=1).font = Font(size=9, italic=True, color="555555")
ws.cell(row=fn, column=1).alignment = Alignment(wrap_text=True, vertical="top")

# column widths / row heights
ws.column_dimensions["A"].width = 16
for col in "BCDEFGHIJK": ws.column_dimensions[col].width = 13
ws.row_dimensions[1].height = 26; ws.row_dimensions[2].height = 20; ws.row_dimensions[3].height = 46
for i in range(len(data)): ws.row_dimensions[4 + i].height = 22
ws.row_dimensions[fn].height = 46
ws.freeze_panes = "B4"

fn_out = os.path.join(OUT, "R1Q5_reproducibility_table.xlsx")
wb.save(fn_out)
print("saved", fn_out)
