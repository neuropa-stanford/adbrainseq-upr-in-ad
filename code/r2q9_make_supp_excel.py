#!/usr/bin/env python3
"""R2.9 — build a Supplementary Excel of the UPR-gene differential expression underlying the
independent-cohort replication (Mathys 2024 + SEA-AD). README + one long-format sheet per cohort."""
import os, glob, gzip, csv, statistics as st
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

OUT = os.path.dirname(os.path.abspath(__file__))
SCR = ("/tmp/"
       "adbrainseq_work")
UPR = set(l.strip() for l in open(os.path.join(OUT, "ERstress_260_geneset.txt")) if l.strip())
SENS = {"EIF2AK3", "ERN1", "ATF6"}
def gs(g): return "UPR-sensor" if g in SENS else "UPR-target(ER-stress)"

# ---- Mathys 2024 (PFC, nft): per broad class, per gene mean logFC_nb + best log10p ----
M = os.path.join(SCR, "mathys24/dereg")
MCL = {"Ast": "Astrocyte", "Exc": "Excitatory neuron", "Inh": "Inhibitory neuron",
       "Mic": "Microglia", "Oli": "Oligodendrocyte", "Opc": "OPC"}
m = {}   # (class, gene) -> [ (logFC, log10p) ]
for fp in glob.glob(os.path.join(M, "aggregated_fullset.*.tsv.gz")):
    pre = os.path.basename(fp).split(".")[1].split("_")[0]
    if pre not in MCL: continue
    with gzip.open(fp, "rt") as fh:
        for row in csv.DictReader(fh, delimiter="\t"):
            if row.get("path") != "nft" or row.get("region") != "PFC": continue
            g = row["gene"]
            if g not in UPR and g not in SENS: continue
            try: fc = float(row["logFC_nb"]); lp = float(row["log10p_nm"])
            except (TypeError, ValueError): continue
            m.setdefault((MCL[pre], g), []).append((fc, lp))

# ---- SEA-AD (MTG, CPS): per broad class, per gene mean logFC + min p ----
SCL = {"Ex": "Excitatory neuron", "In": "Inhibitory neuron", "Ast": "Astrocyte",
       "Mic": "Microglia", "Oli": "Oligodendrocyte", "OPC": "OPC"}
s = {}
for fp in glob.glob(os.path.join(SCR, "seaad", "*.csv")):
    c0 = os.path.basename(fp).split("__")[0]
    if c0 not in SCL: continue
    for row in csv.reader(open(fp)):
        if len(row) < 22: continue
        g = row[0]
        if g not in UPR and g not in SENS: continue
        try: fc = float(row[7]); p = float(row[21])
        except (TypeError, ValueError): continue
        s.setdefault((SCL[c0], g), []).append((fc, p))

wb = openpyxl.Workbook()
BOLD = Font(bold=True); HDRFILL = PatternFill("solid", fgColor="DDE7F0")
def sheet(name):
    ws = wb.create_sheet(name); return ws
def header(ws, cols):
    ws.append(cols)
    for c in ws[1]:
        c.font = BOLD; c.fill = HDRFILL

# README
rd = wb.active; rd.title = "README"
for line in [
    ["Supplementary Data — UPR-associated gene differential expression in two independent AD snRNA cohorts (Reviewer 2, comment 9)"],
    [""],
    ["Cohort", "Source", "Consortium / region", "Donors", "Braak stages", "DE metric"],
    ["Mathys 2024", "Mathys H. et al., Nature 2024 (Suppl. Table 9)", "ROSMAP / prefrontal cortex (PFC subset)", "48",
     "Braak I–VI", "per-cell-type differential expression vs NFT-tangle burden (path=nft); we use logFC_nb (negative-binomial "
     "model coefficient); a MAST coefficient (coef_mast) and neg-log10 p (log10p_nm) are also provided in the source table. sign = direction."],
    ["SEA-AD 2024", "Gabitto M. et al., Nat. Neurosci. 2024; sea-ad-single-cell-profiling (AWS open data)",
     "Allen Institute / middle temporal gyrus (MTG) — INDEPENDENT of ROSMAP", "84",
     "Braak 0–VI (0:2, II:4, III:6, IV:23, V:34, VI:15)",
     "NEBULA negative-binomial MIXED model on raw UMI counts (library-size offset = depth normalization); donor = random "
     "effect (donor-level, not pseudoreplicated); expression regressed on Continuous Pseudo-progression Score (CPS), "
     "adjusted for age, sex, genes-detected, race, 10x method. logFC = NB coefficient per unit CPS (natural-log scale); sign = direction."],
    [""],
    ["SEA-AD (MTG) cohort detail — 84 donors, analysed as a CONTINUOUS neuropathology axis (not a binary control/AD split):"],
    ["  Source studies", "ACT (Adult Changes in Thought) = 69 ; ADRC Clinical Core = 15"],
    ["  Braak (NFT)", "0: 2 | II: 4 | III: 6 | IV: 23 | V: 34 | VI: 15"],
    ["  Thal phase (amyloid)", "0: 9 | 1: 5 | 2: 7 | 3: 12 | 4: 30 | 5: 21"],
    ["  ADNC (overall AD neuropath. change)", "Not AD: 9 | Low: 12 | Intermediate: 21 | High: 42"],
    ["  Cognitive status", "Dementia: 42 | No dementia: 42 (balanced)"],
    ["  Sex / APOE4 carriers / Age", "Female 51, Male 33 | APOE4+ = 25/84 | age at death 65-102 (mean 89)"],
    ["  Reference / low-pathology anchor", "ADNC 'Not AD' (n=9) and Braak 0-II (n=6) anchor the low end; DE is regressed on the"],
    ["    ", "Continuous Pseudo-progression Score (CPS) across all 84 donors (adjusted for age, sex, genes-detected, method)."],
    ["Mathys 2024 (ROSMAP, PFC): 48 donors spanning Braak I-VI; DE regressed on NFT-tangle burden (continuous)."],
    [""],
    ["Notes:"],
    ["- logFC < 0 = transcript decreases with increasing pathology (down);  logFC > 0 = increases (up)."],
    ["- Values are aggregated to broad cell classes as the mean logFC across the cohort's fine subtypes (n_subtypes given)."],
    ["- Gene set: 'UPR-target(ER-stress)' = curated Response-to-ER-stress set (GO:0034976, 260 genes); 'UPR-sensor' = EIF2AK3/ERN1/ATF6."],
    ["- Region: discovery = Mathys 2019 ROSMAP prefrontal cortex (BA10); Mathys 2024 multi-region, PFC subset used here (matched); SEA-AD = MTG."],
    ["- Independence: Mathys 2024 draws from the SAME ROSMAP biobank as the discovery cohort; exact donor overlap could NOT be verified"],
    ["    from available IDs (incompatible schemes) but is plausible -> Mathys 2024 is NOT fully independent. SEA-AD (Allen, 84 donors, MTG) IS fully independent."],
    ["- Result: UPR-target transcripts decrease in excitatory & inhibitory neurons in BOTH cohorts; increase in oligodendrocytes."],
]:
    rd.append(line)
rd["A1"].font = Font(bold=True, size=12)
for c in rd[3]:
    c.font = BOLD; c.fill = HDRFILL

# Mathys sheet
ws = sheet("Mathys2024_PFC_vs_tangles")
header(ws, ["cell_type", "gene", "gene_set", "logFC_vs_tangles(mean)", "neg_log10p(best)", "n_subtypes"])
for (cls, g), vals in sorted(m.items()):
    ws.append([cls, g, gs(g), round(st.mean([v[0] for v in vals]), 4),
               round(max(v[1] for v in vals), 2), len(vals)])
# SEA-AD sheet
ws = sheet("SEA-AD_MTG_vs_CPS")
header(ws, ["cell_type", "gene", "gene_set", "logFC_vs_CPS(mean)", "p_CPS(min)", "n_subtypes"])
for (cls, g), vals in sorted(s.items()):
    ws.append([cls, g, gs(g), round(st.mean([v[0] for v in vals]), 4),
               f"{min(v[1] for v in vals):.2e}", len(vals)])
for wsx in wb.worksheets:
    for col in wsx.columns:
        w = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        wsx.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 60)

fn = os.path.join(os.path.dirname(OUT), "R2Q9_Supplementary_DEG_UPR_replication.xlsx")
wb.save(fn)
print("saved", os.path.basename(fn))
print(f"  Mathys2024 rows: {len(m)} | SEA-AD rows: {len(s)}")
