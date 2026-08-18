#!/usr/bin/env python3
"""R1.1 (Jonathan's request) — UPR GO-term matrix.

18 GO terms taken from the reference figure panels a/b, tested in
Thapsigargin (positive control) / Mizuno / Nativio, with and without a
p<0.05 DEG cutoff, split into UP / DOWN / BOTH.

Enrichment: g:Profiler gost API (same tool the manuscript used), GO:BP,
g:SCS multiple-testing correction, default (whole-genome) background,
all_results=True so non-significant terms are also reported.
Inputs = SuppD1/D2/D3 from the ANC submission (our own files).
"""
import json, os, math, urllib.request, sys
import openpyxl

BASE = "/data/adbrainseq/anc"
OUT = "/data/adbrainseq/Stanford U/ClaudeAgentwithWIKILLM/subprojects/adbrainseq-manuscript/R1Q1_figure_data"
CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "gost_cache")
os.makedirs(CACHE, exist_ok=True)

# ---------------------------------------------------------------- GO terms
GO_TERMS = [
    ("GO:0034976", "Response to ER stress", "a"),
    ("GO:0006986", "Response to unfolded protein", "a"),
    ("GO:0035966", "Response to topologically incorrect protein", "a"),
    ("GO:0035967", "Cellular response to topologically incorrect protein", "a"),
    ("GO:0034620", "Cellular response to unfolded protein", "a"),
    ("GO:0030968", "ER UPR", "a"),
    ("GO:0036498", "IRE1-mediated UPR", "a"),
    ("GO:1905897", "Regulation of response to ER stress", "a"),
    ("GO:1903573", "Neg. regulation of response to ER stress", "a"),
    ("GO:0036503", "ERAD pathway", "a"),
    ("GO:0006457", "Protein folding", "a"),
    ("GO:0034975", "Protein folding in ER", "a"),
    ("GO:0006888", "ER to Golgi vesicle-mediated transport", "a"),
    ("GO:0042886", "Amide transport", "a"),
    ("GO:0030433", "Ubiquitin-dependent ERAD pathway", "b"),
    ("GO:0008104", "Protein localization", "b"),
    ("GO:0015031", "Protein transport", "b"),
    ("GO:0015833", "Peptide transport", "b"),
]
GO_IDS = [g[0] for g in GO_TERMS]

# ---------------------------------------------------------------- load data
def clean(s):
    return str(s).strip() if s is not None else ""

def load_thap():
    """SuppD1 — thapsigargin DESeq2. symbol, log2FC, DESeq2 p."""
    wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD1_Original Thap_DESeq.xlsx"), read_only=True)
    ws = wb.active
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        sym = clean(r[2])
        try:
            lfc = float(r[3]); p = float(r[5])
        except (TypeError, ValueError):
            continue
        if not sym or sym == "None":
            continue
        if sym not in out or p < out[sym][1]:
            out[sym] = (lfc, p)
    wb.close()
    return out

def load_mizuno():
    """SuppD2 — Mizuno. A=symbol, I=log2FC(AD vs non-AD), L=P value."""
    wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD2_MizunoSeq.xlsx"), read_only=True)
    ws = wb.active
    out = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        sym = clean(r[0])
        try:
            lfc = float(r[8]); p = float(r[11])
        except (TypeError, ValueError):
            continue
        if not sym or sym == "None":
            continue
        if sym not in out or p < out[sym][1]:
            out[sym] = (lfc, p)
    wb.close()
    return out

def welch(a, b):
    na, nb = len(a), len(b)
    if na < 2 or nb < 2:
        return None
    ma, mb = sum(a) / na, sum(b) / nb
    va = sum((x - ma) ** 2 for x in a) / (na - 1)
    vb = sum((x - mb) ** 2 for x in b) / (nb - 1)
    se2 = va / na + vb / nb
    if se2 <= 0:
        return None
    t = (ma - mb) / math.sqrt(se2)
    df = se2 ** 2 / ((va / na) ** 2 / (na - 1) + (vb / nb) ** 2 / (nb - 1))
    # two-sided p from Student t via incomplete beta
    x = df / (df + t * t)
    p = betainc(df / 2.0, 0.5, x)
    return ma, mb, t, max(min(p, 1.0), 0.0)

def betacf(a, b, x):
    MAXIT, EPS, FPMIN = 200, 3e-16, 1e-300
    qab, qap, qam = a + b, a + 1.0, a - 1.0
    c, d = 1.0, 1.0 - qab * x / qap
    if abs(d) < FPMIN: d = FPMIN
    d = 1.0 / d; h = d
    for m in range(1, MAXIT + 1):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d; d = FPMIN if abs(d) < FPMIN else d
        c = 1.0 + aa / c; c = FPMIN if abs(c) < FPMIN else c
        d = 1.0 / d; h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d; d = FPMIN if abs(d) < FPMIN else d
        c = 1.0 + aa / c; c = FPMIN if abs(c) < FPMIN else c
        d = 1.0 / d; de = d * c; h *= de
        if abs(de - 1.0) < EPS: break
    return h

def betainc(a, b, x):
    if x <= 0: return 0.0
    if x >= 1: return 1.0
    lb = (math.lgamma(a + b) - math.lgamma(a) - math.lgamma(b)
          + a * math.log(x) + b * math.log(1 - x))
    if x < (a + 1) / (a + b + 2):
        return math.exp(lb) * betacf(a, b, x) / a
    return 1.0 - math.exp(lb) * betacf(b, a, 1 - x) / b

def load_nativio():
    """SuppD3 — Nativio raw counts; 12 AD (G:R) vs 10 old (S:AB).
    log2FC = log2(mean_AD/mean_old); p = Welch t-test on counts.
    (SuppD3 contains no p-values; DESeq2 reprocessing is a separate TODO.)"""
    wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD3_NativioSeq.xlsx"), read_only=True)
    ws = wb.active
    out = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        sym = clean(r[1])
        if not sym or sym == "None":
            continue
        try:
            ad = [float(x) for x in r[6:18] if x is not None]
            old = [float(x) for x in r[18:28] if x is not None]
        except (TypeError, ValueError):
            continue
        w = welch(ad, old)
        if w is None:
            continue
        ma, mb, t, p = w
        if ma <= 0 or mb <= 0:
            continue
        lfc = math.log2(ma / mb)
        if sym not in out or p < out[sym][1]:
            out[sym] = (lfc, p)
    wb.close()
    return out

# ---------------------------------------------------------------- gost
def gost(genes, tag):
    fn = os.path.join(CACHE, tag + ".json")
    if os.path.exists(fn):
        return json.load(open(fn))
    payload = {"organism": "hsapiens", "query": sorted(genes), "sources": ["GO:BP"],
               "all_results": True, "no_evidences": True, "user_threshold": 0.05,
               "significance_threshold_method": "g_SCS"}
    req = urllib.request.Request("https://biit.cs.ut.ee/gprofiler/api/gost/profile/",
                                 data=json.dumps(payload).encode(),
                                 headers={"Content-Type": "application/json"})
    res = json.load(urllib.request.urlopen(req, timeout=900))["result"]
    keep = [r for r in res if r["native"] in GO_IDS]
    json.dump(keep, open(fn, "w"))
    return keep

# ---------------------------------------------------------------- main
def main():
    data = {"Thapsigargin": load_thap(), "Mizuno": load_mizuno(), "Nativio": load_nativio()}
    for k, v in data.items():
        print(k, len(v), "genes", file=sys.stderr)

    rows = []
    for ds, d in data.items():
        for thr, cut in [("no cutoff (sign only)", None), ("p<0.05", 0.05)]:
            sel = {g: v for g, v in d.items() if cut is None or v[1] < cut}
            up = [g for g, v in sel.items() if v[0] > 0]
            dn = [g for g, v in sel.items() if v[0] < 0]
            both = up + dn
            for direction, genes in [("UP", up), ("DOWN", dn), ("BOTH", both)]:
                tag = f"{ds}__{'nocut' if cut is None else 'p05'}__{direction}"
                print("query", tag, len(genes), file=sys.stderr)
                if not genes:
                    res = []
                else:
                    res = gost(genes, tag)
                by = {r["native"]: r for r in res}
                for go, name, panel in GO_TERMS:
                    r = by.get(go)
                    rows.append({
                        "dataset": ds, "threshold": thr, "direction": direction,
                        "n_query_genes": len(genes), "GO": go, "term": name,
                        "figure_panel": panel,
                        "adjP": r["p_value"] if r else "",
                        "neglog10_adjP": (-math.log10(r["p_value"]) if r and r["p_value"] > 0 else
                                          (330 if r else "")),
                        "genes_in_term": r["intersection_size"] if r else "",
                        "term_size": r["term_size"] if r else "",
                        "significant": (r["p_value"] < 0.05) if r else False,
                    })

    import csv
    fn = os.path.join(OUT, "R1Q1_UPR_GO_matrix_18terms_up_down_both.csv")
    with open(fn, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        w.writeheader(); w.writerows(rows)
    print("wrote", fn, len(rows), "rows", file=sys.stderr)

if __name__ == "__main__":
    main()
