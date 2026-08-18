#!/usr/bin/env python3
"""R2.6 — specificity of the UPR change against the global transcriptional shift.

Reviewer 2 asked for: expression-matched random gene sets, negative-control pathways,
adjustment for library complexity, comparison against the global shift, and a competitive
(rather than self-contained) enrichment test.

Two complementary tests, both run on the same rank metric used for GSEA:

1. EXPRESSION-MATCHED EMPIRICAL NULL — 10,000 random gene sets matched to each real set
   for size AND expression decile composition. Answers "does this set shift more than
   equally-expressed random genes?" (controls abundance/detection bias).

2. cameraPR — competitive gene-set test (limma formulation): the set's statistics are
   compared against ALL OTHER genes with a variance inflation factor
   VIF = 1 + (K-1)*inter-gene correlation, so the comparison is explicitly relative to the
   genome-wide background rather than to zero.

Negative-control pathways (biologically unrelated) are run through the identical pipeline.
"""
import csv, math, os, sys
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from r1q1_gomatrix import load_mizuno, GO_TERMS, BASE, welch, clean  # noqa
from r1q1_gsea import gene_set, load_nativio_t  # noqa
import openpyxl

OUT = ("/data/adbrainseq/Publication/2024_scRNA seq/"
       "V1 Science manuscript/ADBrainSeq_V6.0/Acta Neuropathologica Communication/"
       "Major Revision/processed raw data")
NPERM = 10000
rng = np.random.default_rng(7)

NEG_CONTROLS = [
    ("GO:0007608", "Sensory perception of smell"),
    ("GO:0006936", "Muscle contraction"),
    ("GO:0140014", "Mitotic nuclear division"),
    ("GO:0006954", "Inflammatory response"),
    ("GO:0007586", "Digestion"),
]
CORE = ["GO:0034976", "GO:0030968", "GO:0006986", "GO:0034620", "GO:0035966",
        "GO:0035967", "GO:1905897", "GO:1903573", "GO:0036503"]
FOLD = ["GO:0006457", "GO:0034975", "GO:0006888"]
GENERIC = ["GO:0008104", "GO:0015031", "GO:0015833"]
NAME = {g: n for g, n, _ in GO_TERMS}
NAME.update(dict(NEG_CONTROLS))

# ---------------------------------------------------------------- input data
def mizuno_with_abundance():
    """SuppD2: A symbol, I log2FC, J average abundance log2(CPM), L P value."""
    wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD2_MizunoSeq.xlsx"), read_only=True)
    ws = wb.active
    out = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        sym = clean(r[0])
        try:
            lfc, ab, p = float(r[8]), float(r[9]), float(r[11])
        except (TypeError, ValueError):
            continue
        if not sym or sym == "None":
            continue
        if sym not in out or p < out[sym][2]:
            out[sym] = (math.copysign(-math.log10(max(p, 1e-300)), lfc), ab, p)
    wb.close()
    return out

def nativio_with_abundance():
    wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD3_NativioSeq.xlsx"), read_only=True)
    ws = wb.active
    out = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        sym = clean(r[1])
        if not sym or sym == "None":
            continue
        try:
            ad = [float(x) for x in r[6:18] if x is not None]
            old = [float(x) for x in r[18:28] if x is not None]
        except (TypeError, ValueError):
            continue
        w = welch(ad, old)
        if w is None:
            continue
        ma, mb, t, p = w
        if ma <= 0 or mb <= 0:
            continue
        ab = math.log2((sum(ad) + sum(old)) / (len(ad) + len(old)) + 1)
        if sym not in out or p < out[sym][2]:
            out[sym] = (t, ab, p)
    wb.close()
    return out

# ---------------------------------------------------------------- tests
def expression_matched_null(stat, decile, members, nperm=NPERM):
    """Observed mean statistic vs random sets matched on size and expression decile."""
    pos = np.array([i for i, g in enumerate(members) if g is not None])
    idx_by_dec = {d: np.where(decile == d)[0] for d in range(10)}
    obs_idx = members
    if len(obs_idx) < 10:
        return None
    obs = stat[obs_idx].mean()
    counts = np.bincount(decile[obs_idx], minlength=10)
    null = np.empty(nperm)
    for b in range(nperm):
        pick = np.concatenate([rng.choice(idx_by_dec[d], c, replace=False)
                               for d, c in enumerate(counts) if c > 0])
        null[b] = stat[pick].mean()
    mu, sd = null.mean(), null.std(ddof=1)
    z = (obs - mu) / sd if sd > 0 else np.nan
    p = (np.sum(np.abs(null - mu) >= abs(obs - mu)) + 1) / (nperm + 1)
    return {"obs_mean_stat": float(obs), "null_mean": float(mu), "null_sd": float(sd),
            "z_vs_matched_null": float(z), "p_matched_null": float(p)}

def camera_pr(stat, members, inter_gene_cor=0.01):
    """limma cameraPR: competitive test of set vs all other genes, VIF-corrected."""
    N = len(stat)
    K = len(members)
    if K < 10 or K >= N:
        return None
    m = np.zeros(N, bool); m[members] = True
    x1, x2 = stat[m], stat[~m]
    n1, n2 = K, N - K
    v1, v2 = x1.var(ddof=1), x2.var(ddof=1)
    sp2 = ((n1 - 1) * v1 + (n2 - 1) * v2) / (n1 + n2 - 2)
    vif = 1 + (n1 - 1) * inter_gene_cor
    se = math.sqrt(sp2 * (vif / n1 + 1 / n2))
    t = (x1.mean() - x2.mean()) / se
    df = n1 + n2 - 2
    from r1q1_gomatrix import betainc
    p = betainc(df / 2.0, 0.5, df / (df + t * t))
    return {"camera_t": float(t), "camera_p": float(min(max(p, 0.0), 1.0)),
            "camera_direction": "Up" if t > 0 else "Down",
            "mean_set_stat": float(x1.mean()), "mean_background_stat": float(x2.mean())}

def bh(ps):
    ps = np.array(ps, float); n = len(ps)
    o = np.argsort(ps)
    adj = np.minimum.accumulate((ps[o] * n / (np.arange(n) + 1))[::-1])[::-1]
    out = np.empty(n); out[o] = np.minimum(adj, 1.0)
    return out

# ---------------------------------------------------------------- main
def main():
    all_ids = CORE + FOLD + GENERIC + [g for g, _ in NEG_CONTROLS]
    sets = {g: set(gene_set(g)) for g in all_ids}
    for g in all_ids:
        print(f"{g} {NAME[g][:44]:46s} {len(sets[g]):5d} genes", file=sys.stderr)

    data = {"Mizuno": mizuno_with_abundance(), "Nativio": nativio_with_abundance()}
    rows = []
    for ds, d in data.items():
        genes = np.array(list(d.keys()))
        stat = np.array([d[g][0] for g in genes], float)
        abund = np.array([d[g][1] for g in genes], float)
        decile = np.clip((np.argsort(np.argsort(abund)) * 10 // len(abund)), 0, 9)
        gi = {g: i for i, g in enumerate(genes)}
        print(f"\n=== {ds}: {len(genes)} genes, global mean stat = {stat.mean():+.3f}",
              file=sys.stderr)

        res = []
        for g in all_ids:
            mem = np.array(sorted(gi[s] for s in sets[g] if s in gi))
            grp = ("UPR/ERAD core" if g in CORE else "folding/trafficking" if g in FOLD
                   else "generic parent" if g in GENERIC else "negative control")
            row = {"dataset": ds, "GO": g, "term": NAME[g], "group": grp,
                   "K_in_data": len(mem), "global_mean_stat": float(stat.mean())}
            if len(mem) < 10:
                row["note"] = "fewer than 10 genes — not tested"
                res.append(row); continue
            em = expression_matched_null(stat, decile, mem)
            cp = camera_pr(stat, mem)
            row.update(em or {}); row.update(cp or {})
            res.append(row)
        tested = [r for r in res if "p_matched_null" in r]
        for r, q1, q2 in zip(tested, bh([r["p_matched_null"] for r in tested]),
                             bh([r["camera_p"] for r in tested])):
            r["padj_matched_null"] = float(q1)
            r["padj_camera"] = float(q2)
            r["specific_vs_background"] = bool(q1 < 0.05 and q2 < 0.05)
        rows += res

    keys = ["dataset", "group", "GO", "term", "K_in_data", "global_mean_stat",
            "obs_mean_stat", "null_mean", "null_sd", "z_vs_matched_null",
            "p_matched_null", "padj_matched_null", "mean_set_stat", "mean_background_stat",
            "camera_t", "camera_direction", "camera_p", "padj_camera",
            "specific_vs_background", "note"]
    fn = os.path.join(OUT, "R2q6_specificity_expression_matched_and_camera.csv")
    with open(fn, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=keys, extrasaction="ignore")
        w.writeheader(); w.writerows(rows)
    print("\nwrote", fn, len(rows), "rows", file=sys.stderr)

    for ds in data:
        print(f"\n===== {ds} =====", file=sys.stderr)
        for r in rows:
            if r["dataset"] != ds or "z_vs_matched_null" not in r:
                continue
            print(f"  [{r['group'][:18]:18s}] {r['term'][:40]:42s} K={r['K_in_data']:4d} "
                  f"z={r['z_vs_matched_null']:+6.2f} BHp={r['padj_matched_null']:.4f} | "
                  f"camera {r['camera_direction']:4s} BHp={r['padj_camera']:.2e} "
                  f"{'SPECIFIC' if r['specific_vs_background'] else ''}", file=sys.stderr)

if __name__ == "__main__":
    main()
