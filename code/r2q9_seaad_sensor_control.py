#!/usr/bin/env python3
"""SEA-AD internal control — the three canonical UPR SENSOR transcripts EIF2AK3(PERK), ERN1(IRE1),
ATF6 across AD stage (Braak grouping: low-Braak 0,I,II / early-AD III,IV / late-AD V,VI),
donor-level. Per cell type, two violins = distribution over donors of (per-donor sensor expression −
mean low-Braak expression) for early-AD and late-AD. Dashed 0 = low-Braak baseline. If sensors are flat,
violins sit on 0. Annotation = donor-level ordinal-Braak trend (Spearman, n=84); RED if p<0.05."""
import os, csv
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams.update({"font.family": "Arial", "font.sans-serif": ["Arial"], "axes.unicode_minus": False, "pdf.fonttype": 42, "ps.fonttype": 42})
from matplotlib.patches import Patch
from scipy.stats import spearmanr

OUT = os.path.dirname(os.path.abspath(__file__))
rows = list(csv.DictReader(open(os.path.join(OUT, "SEAAD_3group_donor_pseudobulk.csv"))))
CELLS = [("Ex", "Excitatory\nneurons"), ("In", "Inhibitory\nneurons"), ("Ast", "Astrocytes"),
         ("Mic", "Microglia"), ("Oli", "Oligodendrocytes"), ("OPC", "Oligodendrocyte\nprogenitor cells")]
SENSORS = [("EIF2AK3", "EIF2AK3 / PERK", "#1f7a8c"), ("ERN1", "ERN1 / IRE1", "#9c2a4e"), ("ATF6", "ATF6", "#3e52a0")]
C_MID, C_HIGH, RED = "#3c8f34", "#c8d94a", "#e2231a"   # (legacy) superseded by per-cell-type CTCOL
# Figure-5 style: colour each cell type by its own hue (both Braak groups share the hue); regular labels, no stats.
CTCOL = {"Ex": "#746fbd", "In": "#c2b7de", "Ast": "#f2a9a0",
         "Mic": "#cdea9a", "Oli": "#9ad4e7", "OPC": "#cbb0e0"}

# donor -> Braak 0-6
import openpyxl
wb = openpyxl.load_workbook(os.path.join(OUT, "SEAAD_donor_metadata_SuppTable1.xlsx"), read_only=True); ws = wb.active
h = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
di = next(i for i, c in enumerate(h) if "donor id" in c.lower()); bi = next(i for i, c in enumerate(h) if c.lower().startswith("braak"))
bmap = {"Braak 0": 0, "Braak I": 1, "Braak II": 2, "Braak III": 3, "Braak IV": 4, "Braak V": 5, "Braak VI": 6}
d2b = {r[di]: bmap.get(str(r[bi])) for r in ws.iter_rows(min_row=2, values_only=True) if r[di]}; wb.close()

def group(b): return "low" if b in (0, 1, 2) else "mid" if b in (3, 4) else "high" if b in (5, 6) else None

def pstar(p): return "****" if p < 1e-4 else "***" if p < 1e-3 else "**" if p < 1e-2 else "*" if p < 0.05 else "ns"

fig, axes = plt.subplots(len(SENSORS), 1, figsize=(12.5, 9.6), sharex=True)
for row_i, (g, glab, gcol) in enumerate(SENSORS):
    ax = axes[row_i]
    ax.axhline(0, color="#000", lw=0.9, ls=":", zorder=1)
    lim = 0.0
    per = {}   # cell -> dict(group->list per-donor relative)
    trends = {}
    for c, _ in CELLS:
        vals = {"low": [], "mid": [], "high": []}; bk_all = []; ex_all = []
        for r in rows:
            if r["cell_type"] != c: continue
            b = d2b.get(r["donor"]); grp = group(b) if b is not None else None
            if grp is None: continue
            e = float(r[g]); vals[grp].append((r["donor"], e)); bk_all.append(b); ex_all.append(e)
        base = np.mean([e for _, e in vals["low"]]) if vals["low"] else 0.0
        rel = {k: np.array([e - base for _, e in vals[k]]) for k in vals}
        per[c] = rel
        for k in ("mid", "high"):
            if len(rel[k]): lim = max(lim, np.abs(rel[k]).max())
        trends[c] = spearmanr(ex_all, bk_all) if len(bk_all) > 3 else (np.nan, np.nan)
    lim = (lim or 0.1) * 1.15; YL = (-lim, lim); bw = 0.34
    for i, (c, lab) in enumerate(CELLS):
        for k, grp in enumerate(["mid", "high"]):
            v = np.clip(per[c][grp], *YL)
            if not len(v): continue
            xpos = i + (k - 0.5) * bw
            parts = ax.violinplot([v], positions=[xpos], widths=bw * 0.92, showmeans=False, showextrema=False)
            for pc in parts["bodies"]:
                pc.set_facecolor(CTCOL[c]); pc.set_alpha(0.62); pc.set_edgecolor("#1f1f1f"); pc.set_linewidth(0.6)
            jit = (np.random.RandomState(i * 2 + k).rand(len(v)) - 0.5) * bw * 0.7
            ax.scatter(xpos + jit, v, s=4, color="#111", alpha=0.5, edgecolor="none", zorder=3)
            ax.hlines(v.mean(), xpos - bw * .48, xpos + bw * .48, color=RED, lw=2.4, zorder=6)
    ax.set_ylim(*YL); ax.set_xlim(-0.6, len(CELLS) - 0.4)
    ax.set_ylabel(f"{glab}\nΔlog$_2$ expr vs low-Braak", fontsize=9.5, fontweight="normal")
    for sp in ("top", "right"): ax.spines[sp].set_visible(False)
    ax.tick_params(labelbottom=(row_i == len(SENSORS) - 1))
_xt = [i + (k - 0.5) * bw for i in range(len(CELLS)) for k in (0, 1)]
axes[-1].set_xticks(_xt); axes[-1].set_xticklabels(["III,IV", "V,VI"] * len(CELLS), fontsize=8)
# cell-type headers below the Braak-group ticks (regular weight, Fig-5 style)
from matplotlib.transforms import blended_transform_factory
tr = blended_transform_factory(axes[-1].transData, axes[-1].transAxes)
for i, (c, lab) in enumerate(CELLS):
    axes[-1].text(i, -0.20, lab, ha="center", va="top", fontsize=9.5, transform=tr, linespacing=0.95)
fig.legend(handles=[plt.Line2D([0], [0], color=RED, lw=2.4, label="group mean"),
                    plt.Line2D([0], [0], marker="o", color="none", markerfacecolor="#111", markersize=4, label="donor")],
           frameon=False, fontsize=8.5, loc="lower center", ncol=2, bbox_to_anchor=(0.5, -0.005))
# titles/subtitles removed -> moved to the Figure 5 legend (publication style)
fig.subplots_adjust(top=0.97, bottom=0.15, left=0.10, right=0.985, hspace=0.18)
for ext in ("png", "svg", "pdf"):
    fig.savefig(os.path.join(OUT, f"R2Q9_SEAAD_sensor_control.{ext}"), dpi=300 if ext == "png" else None,
                facecolor="white", bbox_inches="tight")
print("saved R2Q9_SEAAD_sensor_control.{png,svg}")
