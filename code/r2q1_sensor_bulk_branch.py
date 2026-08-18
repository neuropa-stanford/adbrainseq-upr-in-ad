#!/usr/bin/env python3
"""BULK sensor-gene expression by BRANCH colour (manuscript palette): EIF2AK3/PERK = cyan,
ERN1/IRE1 = maroon, ATF6 = blue. Rows = branch, cols = Thapsigargin / Mizuno / Nativio.
Per-sample violins: control (grey) vs disease (branch colour). n = each sample. Two-sided
Mann-Whitney U. Sensor transcript level is not a readout of activation (Reviewer 2, comment 1)."""
import os, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams.update({"font.family": "Arial", "font.sans-serif": ["Arial"], "axes.unicode_minus": False})
import openpyxl
from scipy.stats import mannwhitneyu

ROOT = "/data/adbrainseq"
BASE = os.path.join(ROOT, "Publication/2024_scRNA seq/V1 Science manuscript/ADBrainSeq_V6.0/"
                    "Acta Neuropathologica Communication")
OUT = os.path.join(BASE, "Major Revision", "processed raw data")
BRANCH = [("EIF2AK3", "PERK", "#5FC7D8", "#1F8EA0"),
          ("ERN1", "IRE1", "#9C2A4E", "#611026"),
          ("ATF6", "ATF6", "#3E52A0", "#232E63")]
GSET = [b[0] for b in BRANCH]

def thap(fn):
    p = os.path.join(ROOT, "Stanford U/PERK BioSensor Bulk RNA Seq Analysis/Data analysis/"
                     "1st DATA from BGI/GEO deposit/geo_submission_Nov/processed data", fn)
    wb = openpyxl.load_workbook(p, read_only=True); ws = wb.active
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    fc = [i for i, c in enumerate(hdr) if c and "fpkm" in str(c).lower()]
    o = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        s = str(r[2]).strip() if r[2] else ""
        if s in GSET: o[s] = [float(r[i]) for i in fc if isinstance(r[i], (int, float))]
    wb.close(); return o
thap_c, thap_t = thap("RNA seq_Control FPKM.xlsx"), thap("RNA seq_Thapsigargin FPKM.xlsx")

mf = os.path.join(ROOT, "Stanford U/PERK Human AD brain analysis/Human AD brain SEQ analysis/"
                  "Alzheimer's brain disease Bulk RNA seq/2021 Mizuno_Human AD brain RNA seq_decreased PERK/"
                  "2021 Mizuno_GSE173955_Table_1_gene_level_expression.xlsx")
wb = openpyxl.load_workbook(mf, read_only=True); ws = wb["2021 Mizuno_DESeq"]
miz_ad, miz_c = {}, {}
for r in ws.iter_rows(min_row=4, values_only=True):
    s = str(r[0]).strip() if r[0] else ""
    if s in GSET:
        miz_ad[s] = [float(x) for x in r[1:9] if isinstance(x, (int, float))]
        miz_c[s] = [float(x) for x in r[12:22] if isinstance(x, (int, float))]
wb.close()

wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD3_NativioSeq.xlsx"), read_only=True); ws = wb.active
tot_ad = np.zeros(12); tot_old = np.zeros(10); nat = {}
for r in ws.iter_rows(min_row=3, values_only=True):
    s = str(r[1]).strip() if r[1] else ""
    ad = [x if isinstance(x, (int, float)) else 0.0 for x in r[6:18]]
    old = [x if isinstance(x, (int, float)) else 0.0 for x in r[18:28]]
    if len(ad) == 12 and len(old) == 10:
        tot_ad += np.array(ad, float); tot_old += np.array(old, float)
        if s in GSET: nat[s] = (np.array(ad, float), np.array(old, float))
wb.close()
nat_ad = {s: list(ad / tot_ad * 1e6) for s, (ad, old) in nat.items()}   # linear CPM
nat_c = {s: list(old / tot_old * 1e6) for s, (ad, old) in nat.items()}

# per gene: dataset -> (control_vals, disease_vals, unit, clab, dlab) ; values LINEAR (relativised below)
DATA = {"Thapsigargin": (thap_c, thap_t, "FPKM", "Control", "Thap"),
        "Mizuno (GSE173955)": (miz_c, miz_ad, "expression level", "non-AD", "AD"),
        "Nativio": (nat_c, nat_ad, "CPM", "Old", "AD")}
ORDER = ["Thapsigargin", "Mizuno (GSE173955)", "Nativio"]

def pstar(p):
    return "****" if p < 1e-4 else "***" if p < 1e-3 else "**" if p < 1e-2 else "*" if p < 0.05 else "ns"

YLIM = (-2.2, 2.2)
fig, axes = plt.subplots(len(BRANCH), len(ORDER), figsize=(9.6, 8.6), sharey="row")
for gi, (gene, alias, fill, edge) in enumerate(BRANCH):
    for di, dname in enumerate(ORDER):
        ax = axes[gi, di]
        cd, dd, unit, clab, dlab = DATA[dname]
        c0 = np.array(cd.get(gene, []), float); d0 = np.array(dd.get(gene, []), float)
        cmean = float(np.mean(c0)) if len(c0) else np.nan                 # control reference
        c = np.log2(c0 / cmean); d = np.log2(d0 / cmean)                  # relative to control (log2FC)
        ax.axhspan(-0.585, 0.585, color="#f4f4f4", zorder=0)
        ax.axhline(0, color="#999", lw=0.9, ls="--", zorder=1)
        parts = ax.violinplot([c, d], positions=[0, 1], widths=0.8, showmeans=False, showextrema=False)
        for pc, col in zip(parts["bodies"], ["#c9c9c9", fill]):
            pc.set_facecolor(col); pc.set_alpha(0.45); pc.set_edgecolor(edge if col == fill else "#8a8a8a"); pc.set_linewidth(1.1)
        for pos, arr, col, ec in [(0, c, "#c9c9c9", "#8a8a8a"), (1, d, fill, edge)]:
            jit = (np.random.RandomState(gi*10+di).rand(len(arr)) - 0.5) * 0.22
            ax.scatter(pos + jit, arr, s=15, color=col, edgecolor="white", linewidth=0.5, zorder=3)
            ax.hlines(np.mean(arr), pos-0.24, pos+0.24, color=ec, lw=1.8, zorder=4)
        p = mannwhitneyu(c, d, alternative="two-sided").pvalue
        ax.text(0.5, YLIM[1]-0.18, pstar(p), ha="center", va="top", fontsize=8.5)
        ax.set_ylim(*YLIM); ax.set_xlim(-.6, 1.6); ax.set_xticks([0, 1])
        ax.set_xticklabels([f"{clab}\n(n={len(c)})", f"{dlab}\n(n={len(d)})"], fontsize=8)
        ax.tick_params(labelsize=7.5, length=0)
        for sp in ("top", "right"): ax.spines[sp].set_visible(False)
        if di == 0: ax.set_ylabel(f"{gene} / {alias}\n\nlog$_2$FC vs control", fontsize=9.2, fontweight="bold", color=edge)
        if gi == 0: ax.set_title(dname, fontsize=10, fontweight="bold", pad=8)

fig.suptitle("UPR-sensor transcripts in bulk RNA-seq — relative to control (by branch)", fontsize=12.5, fontweight="bold", y=1.01)
fig.text(0.5, 0.972, "EIF2AK3/PERK · ERN1/IRE1 · ATF6 — each point = one sample, log$_2$FC vs the control group mean · bar = mean · "
         "dashed line = no change · grey band = ±1.5-fold · two-sided Mann–Whitney U. Sensor mRNA level is not a measure of sensor activation (Reviewer 2, comment 1).",
         ha="center", va="top", fontsize=7.6, color="#555")
fig.subplots_adjust(top=0.90, bottom=0.055, left=0.10, right=0.985, hspace=0.42, wspace=0.22)
for ext in ("png", "svg"):
    fig.savefig(os.path.join(OUT, f"R2Q1_sensor_bulk_branch.{ext}"), dpi=300 if ext == "png" else None,
                facecolor="white", bbox_inches="tight")
print("saved R2Q1_sensor_bulk_branch.{png,svg}")
