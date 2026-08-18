#!/usr/bin/env python3
"""snRNA sensor-gene expression (EIF2AK3/PERK, ERN1/IRE1, ATF6), DONOR-LEVEL, from the actual
Mathys 2019 count matrix via the Lin-lab team pipeline (Mathys_UMAP_Will_V7.ipynb).

Faithful to the team notebook's normalization (sc.pp.normalize_total target_sum=1e4 then log1p,
cells 34-35) and its per-donor aggregation (groupby(['Subject'..]).mean(), cell 130), but the unit
of inference is the DONOR (Reviewer 2, comment 2 = pseudoreplication fix): we do NOT test across
cells. Per-donor mean log-normalized expression per cell type -> violins by Braak group; n = donors.

Inputs (all local; no Sherlock / no external egress):
  filtered_count_matrix.mtx           17926 genes x 70634 cells (Mathys filtered)
  filtered_gene_row_names.txt         gene order (rows)
  filtered_column_metadata.txt        cell order (cols): TAG, projid, broad.cell.type
  ROSMAP_clinical...xlsx              projid -> braaksc
No pandas (env has a numpy2/pandas ABI clash) -> numpy + scipy + openpyxl + csv only."""
import os, csv, math
import numpy as np
from scipy.io import mmread
from scipy.stats import mannwhitneyu, kruskal
import openpyxl

MATH = ("/data/adbrainseq/Stanford U/Collaboration support/"
        "Prof. Eun-hye Joe Ajou/2019 Mathys/Gene Expression (RNA seq)")
CLIN = ("/data/adbrainseq/Stanford U/PERK Seth Genetics Project/"
        "ROSMAP_clinical_PMI included for Seth.xlsx")
OUT = ("/data/adbrainseq/Publication/2024_scRNA seq/V1 Science manuscript/"
       "ADBrainSeq_V6.0/Acta Neuropathologica Communication/Major Revision/processed raw data")
SENSORS = [("EIF2AK3", "PERK"), ("ERN1", "IRE1"), ("ATF6", "ATF6")]

# ---- gene row indices ----
genes = [l.strip() for l in open(os.path.join(MATH, "filtered_gene_row_names.txt"))]
gidx = {g: i for i, g in enumerate(genes)}
assert all(s in gidx for s, _ in SENSORS), "sensor gene missing"

# ---- cell metadata (order = matrix columns) ----
cell_ct, cell_proj = [], []
with open(os.path.join(MATH, "filtered_column_metadata.txt")) as fh:
    rd = csv.reader(fh, delimiter="\t"); hdr = next(rd)
    ci_ct = hdr.index("broad.cell.type"); ci_pj = hdr.index("projid")
    for row in rd:
        cell_ct.append(row[ci_ct]); cell_proj.append(row[ci_pj])
cell_ct = np.array(cell_ct); cell_proj = np.array(cell_proj)
ncells = len(cell_ct)

# ---- projid -> braaksc ----
wb = openpyxl.load_workbook(CLIN, read_only=True); ws = wb.active
h = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
pj_c, bk_c = h.index("projid"), h.index("braaksc")
proj_braak = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if r[pj_c] is None or r[bk_c] is None: continue
    try:
        proj_braak[str(int(r[pj_c]))] = int(r[bk_c])
    except (TypeError, ValueError):
        continue   # braaksc == 'NA'
wb.close()

# ---- load matrix, normalize per cell (total 1e4, log1p) ----
print("reading matrix (~1.5GB) ...", flush=True)
M = mmread(os.path.join(MATH, "filtered_count_matrix.mtx")).tocsr()   # genes x cells
assert M.shape == (len(genes), ncells), f"shape {M.shape} != ({len(genes)},{ncells})"
colsum = np.asarray(M.sum(axis=0)).ravel().astype(float)              # per-cell library size
colsum[colsum == 0] = 1.0

def norm_lin(gene):
    v = np.asarray(M.getrow(gidx[gene]).todense()).ravel().astype(float)
    return v / colsum * 1e4                                           # normalize_total(1e4), LINEAR

# ---- per-donor mean per cell type ----
donors = sorted(set(cell_proj) & set(proj_braak))
braak = {d: proj_braak[d] for d in donors}
def grp(b): return "low" if b <= 2 else ("int" if b <= 4 else "late")
CELLS = [("Ex", "Excitatory\nneurons"), ("In", "Inhibitory\nneurons"), ("Ast", "Astrocytes"),
         ("Mic", "Microglia"), ("Oli", "Oligodendrocytes"), ("Opc", "OPCs")]

# donor_means[gene][ct] = {donor: mean linear-normalized expression}
donor_means = {s: {ct: {} for ct, _ in CELLS} for s, _ in SENSORS}
for gene, _ in SENSORS:
    le = norm_lin(gene)
    for ct, _ in CELLS:
        cmask = (cell_ct == ct)
        for d in donors:
            m = cmask & (cell_proj == d)
            n = int(m.sum())
            if n >= 10:                      # min-nucleus QC per donor/cell-type (reliable pseudobulk)
                donor_means[gene][ct][d] = float(le[m].mean())

def cohend(a, b):
    a, b = np.asarray(a), np.asarray(b); na, nb = len(a), len(b)
    if na < 2 or nb < 2: return float("nan")
    sp = math.sqrt(((na-1)*a.var(ddof=1)+(nb-1)*b.var(ddof=1))/(na+nb-2))
    return (a.mean()-b.mean())/sp if sp > 0 else float("nan")

# ---- log2FC vs low-Braak (I-II) donor-level reference, per gene x cell type ----
# log2FC vs the Braak I-II (low) reference. Reference = arithmetic mean of DETECTED (v>0) low-group
# donor means (no imputation). Donors where the gene is undetected (v==0) get log2FC = NaN and are
# shown as 'not detected' floor markers by the plotter (so every donor is still displayed).
log2fc = {s: {ct: {} for ct, _ in CELLS} for s, _ in SENSORS}
for gene, _ in SENSORS:
    for ct, _ in CELLS:
        dm = donor_means[gene][ct]
        low_det = [dm[d] for d in dm if grp(braak[d]) == "low" and dm[d] > 0]
        ref = float(np.mean(low_det)) if low_det else float("nan")
        if not (ref and math.isfinite(ref) and ref > 0): continue
        for d, v in dm.items():
            log2fc[gene][ct][d] = math.log2(v / ref) if v > 0 else float("nan")

# ---- save table + print stats (on log2FC) ----
rows = []
for gene, alias in SENSORS:
    for ct, _ in CELLS:
        dm = donor_means[gene][ct]
        fc = log2fc[gene][ct]
        # tests on LINEAR donor means (all donors incl 0) -> rank-based, includes undetected donors
        gl = {k: [dm[d] for d in dm if grp(braak[d]) == k] for k in ("low", "int", "late")}
        # display log2FC (detected donors only) for the reported mean
        gf = {k: [fc[d] for d in fc if grp(braak[d]) == k and np.isfinite(fc[d])] for k in ("low", "int", "late")}
        if all(len(gl[k]) >= 3 for k in ("low", "int", "late")):
            kw = kruskal(gl["low"], gl["int"], gl["late"]).pvalue
            rows.append([gene, alias, ct, "Kruskal-Wallis (3 groups)", len(gl["low"]),
                         f"{len(gl['int'])}/{len(gl['late'])}", "", "", f"{kw:.3g}"])
        for comp in ("int", "late"):
            a, b = np.array(gl[comp]), np.array(gl["low"])
            if len(a) < 3 or len(b) < 3: continue
            p = mannwhitneyu(a, b, alternative="two-sided").pvalue
            mfc = round(float(np.mean(gf[comp])), 3) if gf[comp] else ""
            rows.append([gene, alias, ct, f"{comp} vs low", len(b), len(a),
                         mfc, round(cohend(a, b), 3), f"{p:.3g}"])
with open(os.path.join(OUT, "R2Q1_snRNA_sensor_donorlevel.csv"), "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(["gene", "alias", "cell_type", "test_or_comparison", "n_low", "n_comp",
                "mean_log2FC_comp", "cohens_d", "p"]); w.writerows(rows)
np.save(os.path.join(OUT, "R2Q1_snRNA_sensor_donormeans.npy"),
        np.array([(g, ct, d, braak[d], grp(braak[d]), donor_means[g][ct][d], log2fc[g][ct].get(d, float("nan")))
                  for g, _ in SENSORS for ct, _ in CELLS
                  for d in donor_means[g][ct]], dtype=object), allow_pickle=True)
print(f"donors used: {len(donors)}  (low={sum(grp(braak[d])=='low' for d in donors)}, "
      f"int={sum(grp(braak[d])=='int' for d in donors)}, late={sum(grp(braak[d])=='late' for d in donors)})")
print("gene     cell  test/comparison             n_low n_cmp  meanLog2FC   d      p")
for r in rows:
    mfc = f"{r[6]:9.3f}" if r[6] != "" else " " * 9
    d = f"{r[7]:6.2f}" if r[7] != "" else " " * 6
    print(f"{r[0]:8s} {r[2]:4s} {r[3]:26s} {str(r[4]):>5} {str(r[5]):>5}  {mfc} {d}  {r[8]}")
print("saved R2Q1_snRNA_sensor_donorlevel.csv + donormeans.npy")
