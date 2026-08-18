#!/usr/bin/env python3
"""Supplementary Table S1 — the 122 concordantly-downregulated GO:0034976 genes.
Query g:Profiler (GO:0034976 membership) intersect with concordant-down (down in both cohorts);
attach per-cohort log2FC + p + branch module; write an editable xlsx."""
import sys, json, os, math, urllib.request
sys.path.insert(0, '.')
from r1q1_gomatrix import load_mizuno, load_nativio
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = ("/data/adbrainseq/Publication/2024_scRNA seq/"
       "V1 Science manuscript/ADBrainSeq_V6.0/Acta Neuropathologica Communication/"
       "Major Revision/processed raw data")

miz = load_mizuno(); nat = load_nativio()
common = set(miz) & set(nat)
cd = sorted([g for g in common if miz[g][0] < 0 and nat[g][0] < 0])  # concordant-down
print(f"concordant-down query size: {len(cd)}", file=sys.stderr)

# ---- g:Profiler query (lean: significant terms only, with evidences); reuse cache if present ----
CACHE = os.path.join(OUT, "_gp_S1_lean.json")
if os.path.exists(CACHE):
    resp = json.load(open(CACHE))
    print("loaded cached g:Profiler response", file=sys.stderr)
else:
    payload = {"organism": "hsapiens", "query": cd, "sources": ["GO:BP"], "all_results": False,
               "no_evidences": False, "user_threshold": 0.05, "significance_threshold_method": "g_SCS"}
    req = urllib.request.Request("https://biit.cs.ut.ee/gprofiler/api/gost/profile/",
                                 data=json.dumps(payload).encode(),
                                 headers={"Content-Type": "application/json"})
    resp = json.load(urllib.request.urlopen(req, timeout=900))
    json.dump(resp, open(CACHE, "w"))

go = next((r for r in resp["result"] if r["native"] == "GO:0034976"), None)
assert go is not None, "GO:0034976 not returned"
print(f"GO:0034976 intersection_size (should be 122): {go['intersection_size']}  p={go['p_value']:.2e}",
      file=sys.stderr)

# ---- map intersections -> gene symbols ----
# intersections[i] aligns to meta ...query_1.ensgs[i] (resolved Ensembl IDs).
# mapping is {input_symbol: ensg}; reverse it to recover symbols.
q1 = resp["meta"]["genes_metadata"]["query"]["query_1"]
ensgs = q1["ensgs"]
rev = {e: sym for sym, elist in q1["mapping"].items() for e in elist}
inter = go["intersections"]
in_term = [rev.get(ensgs[i]) for i, ev in enumerate(inter) if ev]
in_term = [g for g in in_term if g and g in miz and g in nat]
print(f"parsed in-term genes: {len(in_term)}", file=sys.stderr)

# validation: the 17 known representative ER-stress genes must all be present
REP = ["CALR","CANX","HSP90B1","PDIA3","P4HB","HYOU1","DNAJB9","SEL1L","DERL1","EDEM3",
       "OS9","AMFR","WFS1","TRIB3","CHAC1","SESN2","EIF2S1"]
missing = [g for g in REP if g not in in_term]
print(f"representative genes missing from parsed set: {missing}", file=sys.stderr)
assert not missing, f"alignment check FAILED, missing reps: {missing}"

genes = sorted(in_term)

# ---- branch module assignment from local curated sets ----
def load_set(fn):
    p = os.path.join(OUT, fn)
    return set(l.strip() for l in open(p) if l.strip()) if os.path.exists(p) else set()
ERAD = load_set("geneset_ERAD.txt"); PERK = load_set("geneset_PERK.txt")
IRE1 = load_set("geneset_IRE1.txt"); ATF6 = load_set("geneset_ATF6.txt")
CHAP = {"CALR","CANX","HSP90B1","HSPA5","PDIA3","P4HB","HYOU1","DNAJB11","PDIA4","PDIA6",
        "SDF2L1","MANF","CRELD2","FKBP11","HSPA13","DNAJC3","DNAJB11"}
# the 17 Panel-C representatives keep their exact figure branch labels (consistency with Fig 1C)
import csv as _csv
REPB = {r["gene"]: r["branch_function"] for r in
        _csv.DictReader(open(os.path.join(OUT, "FIG_representative_conserved_UPR_genes.csv")))}
def branch(g):
    if g in REPB: return REPB[g]              # exact Fig 1C label for the representatives
    if g in ERAD: return "ERAD"
    if g in PERK: return "PERK/ISR"
    if g in IRE1: return "IRE1"
    if g in ATF6: return "ATF6"
    if g in CHAP: return "chaperone/folding"
    return "other UPR/ER-stress"

rows = []
for g in genes:
    lm, pm = miz[g]; ln, pn = nat[g]
    rows.append({"gene": g, "branch_module": branch(g),
                 "Mizuno_log2FC": round(lm, 3), "Mizuno_p": pm,
                 "Nativio_log2FC": round(ln, 3), "Nativio_p": pn,
                 "sig_both_cohorts": "yes" if (pm < 0.05 and pn < 0.05) else "no",
                 "mean_log2FC": round((lm + ln) / 2, 3)})
# order: branch bucket, then most-down first
def bucket(b):
    b = b.lower()
    if "chap" in b or "pdi" in b: return 0
    if "erad" in b: return 1
    if "ire1" in b: return 2
    if "atf6" in b: return 3
    if "perk" in b or "isr" in b: return 4
    return 5
rows.sort(key=lambda r: (bucket(r["branch_module"]), r["mean_log2FC"]))

# ---- write xlsx ----
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "S1_conserved_122"
hdr = ["Gene", "UPR branch module", "Mizuno log2FC", "Mizuno p",
       "Nativio log2FC", "Nativio p", "Significant in both cohorts (p<0.05)", "Mean log2FC"]
head_fill = PatternFill("solid", fgColor="1b2a3a"); head_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="cccccc"); bord = Border(left=thin, right=thin, top=thin, bottom=thin)
ws.append(hdr)
for c in range(1, len(hdr) + 1):
    cell = ws.cell(1, c); cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = bord
for r in rows:
    ws.append([r["gene"], r["branch_module"], r["Mizuno_log2FC"],
               float(f"{r['Mizuno_p']:.3g}"), r["Nativio_log2FC"], float(f"{r['Nativio_p']:.3g}"),
               r["sig_both_cohorts"], r["mean_log2FC"]])
# style body
for ri in range(2, len(rows) + 2):
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(ri, c); cell.border = bord
        if c in (1, 2, 7): cell.alignment = Alignment(horizontal="center" if c != 2 else "left")
        else: cell.alignment = Alignment(horizontal="center")
    ws.cell(ri, 1).font = Font(bold=True)
    if ws.cell(ri, 7).value == "yes":
        ws.cell(ri, 7).fill = PatternFill("solid", fgColor="d6ecd6")
widths = [12, 20, 13, 11, 13, 11, 30, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"
# note row
ws.append([])
ws.append(["Note: 122 genes annotated to GO:0034976 (response to ER stress) that are concordantly "
           "downregulated (negative log2FC) in BOTH the Mizuno (GSE173955) and Nativio (GSE159699) "
           "cohorts. Most show modest per-gene effects; the conserved signal is the reproducible "
           "downward direction across two independent cohorts, not individual significance."])
ws.cell(ws.max_row, 1).font = Font(italic=True, size=9, color="555555")

fn = os.path.join(OUT, "R1Q1_Supplementary_Table_S1_122_conserved_genes.xlsx")
wb.save(fn)
n_sig = sum(1 for r in rows if r["sig_both_cohorts"] == "yes")
print(f"\nSAVED {fn}", file=sys.stderr)
print(f"rows: {len(rows)}  | significant in both cohorts: {n_sig}/{len(rows)}", file=sys.stderr)
from collections import Counter
print("by branch:", dict(Counter(r['branch_module'] for r in rows)), file=sys.stderr)
