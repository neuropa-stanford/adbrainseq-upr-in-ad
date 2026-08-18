#!/usr/bin/env python3
"""R2.9 second independent cohort — Morabito 2021 (GSE174367, human AD PFC snRNA), DONOR-LEVEL.
Per-cell UPR-associated gene expression (union of the PERK/IRE1/ATF6 filtered feature sets) is
aggregated per donor (SampleID) x cell type; AD (n=11) vs Control (n=7) is tested with the DONOR as
the unit (Mann-Whitney, Cohen's d). Direction per cell type checks neuron-down / glia-up.
Control reference = UPR-sensor mRNA (EIF2AK3; only sensor present in the extracted sheets)."""
import os, math, statistics as st
import numpy as np
import openpyxl
from scipy.stats import mannwhitneyu

MOR = ("/data/adbrainseq/Stanford U/PERK Human AD brain analysis/"
       "Human AD brain SEQ analysis/Single cell RNA seq/2021 Morabito_single cell nuclei of AD_Christine/"
       "Morabito_ER_Stress_Filtered_Feature_Dataset_CL_06202022.xlsx")
OUT = os.path.dirname(os.path.abspath(__file__))
SHEETS = ["PERK Filtered Dataset (matrixa)", "IRE1 Filtered Dataset(matrixad)", "ATF6 Filtered Dataset (matrix)"]
SENSORS = {"EIF2AK3", "ERN1", "ATF6"}

wb = openpyxl.load_workbook(MOR, read_only=True)
cells = {}   # barcode -> dict(meta + gene:val)
for sh in SHEETS:
    ws = wb[sh]; h = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    bi, si, di, ci = h.index("Barcode"), h.index("SampleID"), h.index("Diagnosis"), h.index("Cell.Type")
    gcols = [(j, h[j]) for j in range(11, len(h)) if h[j]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        bc = r[bi]
        if bc is None: continue
        d = cells.setdefault(bc, {"sid": r[si], "dx": r[di], "ct": r[ci], "g": {}})
        for j, g in gcols:
            try: d["g"][g] = float(r[j])
            except (TypeError, ValueError): pass
wb.close()

TARGET = sorted({g for c in cells.values() for g in c["g"]} - SENSORS)
NEUR = {"EX", "INH"}; GLIA = {"ASC", "MG", "ODC", "OPC"}
CT_ORDER = ["EX", "INH", "ASC", "MG", "ODC", "OPC"]
CTLAB = {"EX": "Ex (neuron)", "INH": "In (neuron)", "ASC": "Ast (glia)", "MG": "Mic (glia)",
         "ODC": "Oli (glia)", "OPC": "OPC (glia)"}

# per cell: UPR-target score = mean over target genes present; sensor score = EIF2AK3
# aggregate per donor x cell type
agg = {}   # (sid, ct) -> {'dx','upr':[per-cell],'sen':[per-cell]}
for c in cells.values():
    tv = [c["g"][g] for g in TARGET if g in c["g"]]
    if not tv: continue
    k = (c["sid"], c["ct"]); a = agg.setdefault(k, {"dx": c["dx"], "upr": [], "sen": []})
    a["upr"].append(st.mean(tv))
    if "EIF2AK3" in c["g"]: a["sen"].append(c["g"]["EIF2AK3"])
donor = {}  # ct -> {'AD':[donor means], 'Control':[...]} for upr and sen
for (sid, ct), a in agg.items():
    if len(a["upr"]) < 10: continue                       # min-nucleus per donor x cell type
    donor.setdefault(ct, {"AD_upr": [], "Control_upr": [], "AD_sen": [], "Control_sen": []})
    donor[ct][f"{a['dx']}_upr"].append(st.mean(a["upr"]))
    if a["sen"]: donor[ct][f"{a['dx']}_sen"].append(st.mean(a["sen"]))

def cohend(a, b):
    a, b = np.array(a), np.array(b)
    if len(a) < 2 or len(b) < 2: return float("nan")
    sp = math.sqrt(((len(a)-1)*a.var(ddof=1)+(len(b)-1)*b.var(ddof=1))/(len(a)+len(b)-2))
    return (a.mean()-b.mean())/sp if sp > 0 else float("nan")

print(f"UPR target genes used: {len(TARGET)}   donors: AD/Control per cell type below")
print(f"{'cell type':12s}{'nAD':>4}{'nCon':>5}{'UPR meanAD':>11}{'meanCon':>9}{'d':>7}{'p':>8}   dir")
rows = []
for ct in CT_ORDER:
    d = donor.get(ct)
    if not d or len(d["AD_upr"]) < 3 or len(d["Control_upr"]) < 3: continue
    a, b = d["AD_upr"], d["Control_upr"]
    p = mannwhitneyu(a, b, alternative="two-sided").pvalue
    diff = st.mean(a) - st.mean(b); dr = "DOWN" if diff < 0 else "UP"
    print(f"{CTLAB[ct]:12s}{len(a):>4}{len(b):>5}{st.mean(a):>11.3f}{st.mean(b):>9.3f}{cohend(a,b):>7.2f}{p:>8.3f}   {dr}")
    rows.append((ct, CTLAB[ct], len(a), len(b), st.mean(a), st.mean(b), cohend(a, b), p, diff))

neur_down = all(r[8] < 0 for r in rows if r[0] in NEUR)
glia_up = all(r[8] > 0 for r in rows if r[0] in GLIA and r[0] != "OPC")
print(f"\nVERDICT: neurons(EX,INH) all down = {neur_down} ; glia(ASC,MG,ODC) all up = {glia_up}")
np.save(os.path.join(OUT, "R2Q9_morabito_donorlevel.npy"), np.array(rows, dtype=object), allow_pickle=True)
print("saved R2Q9_morabito_donorlevel.npy")
