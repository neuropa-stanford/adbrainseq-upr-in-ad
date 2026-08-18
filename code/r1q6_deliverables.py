#!/usr/bin/env python3
"""R1.6 deliverables: (1) editable Excel supplementary table of the complete Braak-correlation
analysis; (2) figure showing TMED2 (glial-positive) and TRIB3 (neuronal-negative) at the extremes."""
import csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
plt.rcParams.update({"font.family": "Arial", "font.sans-serif": ["Arial"], "axes.unicode_minus": False})

OUT = ("/data/adbrainseq/Publication/2024_scRNA seq/"
       "V1 Science manuscript/ADBrainSeq_V6.0/Acta Neuropathologica Communication/Major Revision/processed raw data")
rows = list(csv.DictReader(open(os.path.join(OUT, "R1Q6_Braak_correlation_UPRgenes.csv"))))
for r in rows:
    r["R"] = float(r["R"]); r["p"] = float(r["p"]); r["adj_p_BH"] = float(r["adj_p_BH"])
CELLS = ["Ex", "In", "Ast", "Mic", "Oli", "Opc"]

# ================= (1) editable Excel supplementary table =================
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Complete correlation"
thin = Side(style="thin", color="C8C8C8"); border = Border(*[thin] * 4)
hdr = ["Cell type", "Gene", "UPR branch", "Pearson R (vs Braak score)", "p value",
       "BH-adjusted p", "n (donors)"]
HF = PatternFill("solid", fgColor="DFE6EC")
for j, h in enumerate(hdr):
    c = ws.cell(row=1, column=j + 1, value=h); c.font = Font(bold=True); c.fill = HF
    c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = border
data = sorted(rows, key=lambda r: (CELLS.index(r["cell_type"]), -r["R"]))
for i, r in enumerate(data):
    vals = [r["cell_type"], r["gene"], r["branch"], round(r["R"], 3), r["p"], r["adj_p_BH"], int(r["n_donors"])]
    for j, v in enumerate(vals):
        c = ws.cell(row=2 + i, column=j + 1, value=v); c.border = border
        c.alignment = Alignment(horizontal="center")
        if r["gene"] in ("TMED2", "TRIB3"):
            c.font = Font(bold=True, color="B00000")
            c.fill = PatternFill("solid", fgColor="FDEEDA")
for col, w in zip("ABCDEFG", [9, 12, 12, 20, 12, 13, 9]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:G{len(data)+1}"

# summary sheet: top 15 pos + 15 neg per cell type
ws2 = wb.create_sheet("Top correlated per cell type")
rr = 1
for ct in CELLS:
    sub = sorted([r for r in rows if r["cell_type"] == ct], key=lambda r: -r["R"])
    ws2.cell(row=rr, column=1, value=f"{ct} — top positively (left) and negatively (right) Braak-correlated UPR genes").font = Font(bold=True, size=11)
    rr += 1
    for h, cc in [("Gene", 1), ("R", 2), ("adj p", 3), ("Gene", 5), ("R", 6), ("adj p", 7)]:
        ws2.cell(row=rr, column=cc, value=h).font = Font(bold=True)
    rr += 1
    top_pos = sub[:15]; top_neg = sub[::-1][:15]
    for k in range(15):
        p, n = top_pos[k], top_neg[k]
        for (g, cc) in [(p["gene"], 1), (round(p["R"], 3), 2), (p["adj_p_BH"], 3),
                        (n["gene"], 5), (round(n["R"], 3), 6), (n["adj_p_BH"], 7)]:
            cell = ws2.cell(row=rr + k, column=cc, value=g)
            if isinstance(g, str) and g in ("TMED2", "TRIB3"):
                cell.font = Font(bold=True, color="B00000")
    rr += 16
wb.save(os.path.join(OUT, "R1Q6_Braak_correlation_supplementary_table.xlsx"))
print("saved xlsx")

# ================= (2) figure: TMED2 / TRIB3 at the extremes =================
panels = [("Oli", "pos", "TMED2"), ("Mic", "pos", "TMED2"), ("In", "neg", "TRIB3")]
fig, axes = plt.subplots(1, 3, figsize=(15, 6.4))
fig.subplots_adjust(left=0.06, right=0.985, top=0.84, bottom=0.08, wspace=0.5)
for ax, (ct, direc, tgt) in zip(axes, panels):
    sub = sorted([r for r in rows if r["cell_type"] == ct], key=lambda r: -r["R"])
    top = sub[:14] if direc == "pos" else sub[::-1][:14]
    if not any(r["gene"] == tgt for r in top):  # ensure target shown
        trec = next(r for r in rows if r["cell_type"] == ct and r["gene"] == tgt)
        top = top[:13] + [trec]
        top = sorted(top, key=lambda r: -r["R"]) if direc == "pos" else sorted(top, key=lambda r: r["R"])
    y = np.arange(len(top))[::-1]
    Rs = [r["R"] for r in top]
    cols = ["#c0392b" if r["gene"] == tgt else ("#e6b0a8" if direc == "neg" else "#7fb3d5") for r in top]
    cols = ["#c0392b" if r["gene"] == tgt else ("#5a9bd4" if direc == "pos" else "#5a9bd4") for r in top]
    cols = ["#b02318" if r["gene"] == tgt else "#9bb7cf" for r in top]
    ax.barh(y, Rs, color=cols, edgecolor="#555", lw=.4, height=.74)
    ax.set_yticks(y)
    ax.set_yticklabels([r["gene"] for r in top], fontsize=8.4)
    for tk, r in zip(ax.get_yticklabels(), top):
        if r["gene"] == tgt: tk.set_color("#b02318"); tk.set_fontweight("bold")
    for yy, r in zip(y, top):
        ax.text(r["R"] + (0.01 if r["R"] >= 0 else -0.01), yy,
                f"{r['R']:+.2f}", va="center", ha="left" if r["R"] >= 0 else "right",
                fontsize=7, color="#b02318" if r["gene"] == tgt else "#555")
    ax.axvline(0, color="#333", lw=.8)
    ax.set_xlabel("Pearson R (expression vs Braak)", fontsize=9)
    for s in ["top", "right"]:
        ax.spines[s].set_visible(False)
    tr = next(r for r in rows if r["cell_type"] == ct and r["gene"] == tgt)
    rank = (sorted([r for r in rows if r["cell_type"] == ct], key=lambda a: -a["R"]).index(tr) + 1
            if direc == "pos" else
            sorted([r for r in rows if r["cell_type"] == ct], key=lambda a: a["R"]).index(tr) + 1)
    tot = sum(1 for r in rows if r["cell_type"] == ct)
    lab = "microglia" if ct == "Mic" else "oligodendrocytes" if ct == "Oli" else \
          "inhibitory neurons" if ct == "In" else "excitatory neurons"
    ax.set_title(f"{lab}\n{tgt}: R={tr['R']:+.2f}, p={tr['p']:.1e}\n"
                 f"rank {rank}/{tot} most {'positive' if direc=='pos' else 'negative'}",
                 fontsize=9.5, fontweight="bold")
fig.suptitle("Selection rationale for TMED2 (glial, Braak-positive) and TRIB3 (neuronal, Braak-negative): "
             "each is at the extreme of the correlation distribution in its cell type",
             fontsize=13, fontweight="bold", y=0.965)
for ext in ("png", "svg"):
    fig.savefig(os.path.join(OUT, f"R1Q6_TMED2_TRIB3_selection.{ext}"),
                dpi=250 if ext == "png" else None, facecolor="white", bbox_inches="tight")
print("saved figure")
