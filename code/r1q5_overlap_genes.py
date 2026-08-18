#!/usr/bin/env python3
"""R1.5 — (1) re-verify the cross-modality reproducibility table counts; (2) export the OVERLAPPED
(reproduced) genes per UPR set with their log2FC in Mizuno, Nativio and snRNA (Ex, In neurons).
Reproduced-DOWN = down in BOTH bulk cohorts AND down in BOTH neuron types (Ex, In)."""
import sys, os, math, csv
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
sys.path.insert(0, ".")
from r1q1_gomatrix import load_mizuno, load_nativio

BASE = ("/data/adbrainseq/Publication/2024_scRNA seq/"
        "V1 Science manuscript/ADBrainSeq_V6.0/Acta Neuropathologica Communication")
OUT = os.path.join(BASE, "Major Revision", "processed raw data")

GS = [("ER-stress (260)", "ERstress_260_geneset.txt"), ("PERK (31)", "geneset_PERK.txt"),
      ("IRE1 (32)", "geneset_IRE1.txt"), ("ATF6 (74)", "geneset_ATF6.txt"), ("ERAD (75)", "geneset_ERAD.txt")]
sets = {n: [l.strip() for l in open(os.path.join(OUT, f)) if l.strip()] for n, f in GS}

miz = {g: v[0] for g, v in load_mizuno().items()}
nat = {g: v[0] for g, v in load_nativio().items()}
# snRNA log2FC (late Braak vs non) per cell type from SuppD6
wb = openpyxl.load_workbook(os.path.join(BASE, "SuppD6_snRNSeqDB.xlsx"), read_only=True); ws = wb.active
CT = {"Ex": 2, "In": 4, "Mic": 8, "Oli": 10}
sn = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    d = {}
    for ct, ci in CT.items():
        try:
            fc = float(r[ci])
            if fc > 0 and math.isfinite(fc): d[ct] = math.log2(fc)
        except (TypeError, ValueError): pass
    if len(d) == 4: sn[str(r[0]).strip()] = d
wb.close()

REP17 = {"CALR","CANX","HSP90B1","PDIA3","P4HB","HYOU1","DNAJB9","SEL1L","DERL1","EDEM3","OS9","AMFR",
         "WFS1","TRIB3","CHAC1","SESN2","EIF2S1"}

# ---- (1) re-verify counts ----
paste = {  # the table the user pasted: set -> (Miz_dn, Nat_dn, bulkOv_dn, snNeuron_dn, repro_dn, Miz_up, Nat_up, bulkOv_up, snGlia_up, repro_up)
 "ER-stress (260)": (109,165,100,177,91,95,39,30,95,13), "PERK (31)": (12,21,12,22,11,15,6,6,15,5),
 "IRE1 (32)": (8,22,7,21,5,20,6,5,20,2), "ATF6 (74)": (32,52,28,59,26,35,15,11,36,7),
 "ERAD (75)": (39,55,37,57,34,23,7,5,37,4)}
print("=== (1) re-verification (recomputed vs pasted) ===", file=sys.stderr)
print(f"{'set':16s}{'Miz↓':>6}{'Nat↓':>6}{'Ov↓':>5}{'snN↓':>6}{'rep↓':>6}{'Miz↑':>6}{'Nat↑':>6}{'Ov↑':>5}{'snG↑':>6}{'rep↑':>6}  match", file=sys.stderr)
gene_rows = []
for name, _ in GS:
    U = [g for g in sets[name] if g in miz and g in nat and g in sn]
    md = sum(1 for g in U if miz[g] < 0); mu = sum(1 for g in U if miz[g] > 0)
    nd = sum(1 for g in U if nat[g] < 0); nu = sum(1 for g in U if nat[g] > 0)
    bd = set(g for g in U if miz[g] < 0 and nat[g] < 0); bu = set(g for g in U if miz[g] > 0 and nat[g] > 0)
    snd = set(g for g in U if sn[g]["Ex"] < 0 and sn[g]["In"] < 0)
    snu = set(g for g in U if sn[g]["Mic"] > 0 and sn[g]["Oli"] > 0)
    ov_d = bd & snd; ov_u = bu & snu
    got = (md, nd, len(bd), len(snd), len(ov_d), mu, nu, len(bu), len(snu), len(ov_u))
    ok = "OK" if got == paste[name] else f"DIFF exp{paste[name]}"
    print(f"{name:16s}" + "".join(f"{x:>6}" for x in got[:5]) + "".join(f"{x:>6}" for x in got[5:]) + f"  {ok}", file=sys.stderr)
    for g in sorted(ov_d, key=lambda g: (miz[g] + nat[g]) / 2):     # reproduced-DOWN genes, most-down first
        gene_rows.append({"gene_set": name, "gene": g,
                          "Mizuno_log2FC": round(miz[g], 3), "Nativio_log2FC": round(nat[g], 3),
                          "snRNA_Ex_log2FC": round(sn[g]["Ex"], 3), "snRNA_In_log2FC": round(sn[g]["In"], 3),
                          "R1.1_representative": "yes" if g in REP17 else ""})

# ---- (2) write overlapped-genes xlsx ----
wb2 = openpyxl.Workbook(); wsx = wb2.active; wsx.title = "reproduced_DOWN_genes"
hdr = ["UPR gene set", "Gene", "Mizuno log2FC", "Nativio log2FC", "snRNA Ex log2FC", "snRNA In log2FC", "R1.1 representative"]
wsx.append(hdr)
hf = PatternFill("solid", fgColor="1b2a3a")
for c in range(1, len(hdr) + 1):
    cell = wsx.cell(1, c); cell.fill = hf; cell.font = Font(bold=True, color="FFFFFF"); cell.alignment = Alignment(horizontal="center", wrap_text=True)
for r in gene_rows:
    wsx.append([r["gene_set"], r["gene"], r["Mizuno_log2FC"], r["Nativio_log2FC"], r["snRNA_Ex_log2FC"], r["snRNA_In_log2FC"], r["R1.1_representative"]])
    if r["R1.1_representative"] == "yes":
        wsx.cell(wsx.max_row, 7).fill = PatternFill("solid", fgColor="d6ecd6")
for i, w in enumerate([16, 10, 13, 13, 15, 15, 18], 1):
    wsx.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
wsx.freeze_panes = "A2"
fn = os.path.join(OUT, "R1Q5_overlapped_reproduced_genes_log2FC.xlsx")
wb2.save(fn)
from collections import Counter
print("\n=== (2) reproduced-DOWN gene table ===", file=sys.stderr)
print("per set:", dict(Counter(r["gene_set"] for r in gene_rows)), file=sys.stderr)
print(f"R1.1 representatives among them: {sum(1 for r in gene_rows if r['R1.1_representative']=='yes')}", file=sys.stderr)
print(f"saved {fn} ({len(gene_rows)} rows)", file=sys.stderr)
